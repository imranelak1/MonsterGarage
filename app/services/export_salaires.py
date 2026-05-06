"""
Export du livre de paie mensuel au format Excel.
Reproduit fidèlement le format du fichier LIVRE DE CAISSE & SALAIRES original.
"""

import calendar
from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.avance_salaire import AvanceSalaire
from app.models.employe import Employe
from app.models.salaire import Salaire

# ── Styles ──────────────────────────────────────────────────────────────────

_JAUNE    = "FFFFD700"
_NOIR     = "FF000000"
_GRIS_CLR = "FF2C3E50"
_GRIS_LGT = "FFF2F2F2"
_BLANC    = "FFFFFFFF"
_ORANGE   = "FFFFA500"
_SLATE_900 = "FF0F172A"
_SLATE_500 = "FF64748B"
_YELLOW_SOFT = "FFFFF7CC"
_LOGO_PATH = Path(__file__).resolve().parent.parent / "static" / "img" / "logo_monster_garage.png"

def _font(bold=False, size=10, color=_NOIR, name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border_thin():
    s = Side(style="thin", color=_NOIR)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom():
    s = Side(style="thin", color=_NOIR)
    return Border(bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _style_range(ws, cell_range: str, *, fill=None, border=None, alignment=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def _add_logo(ws):
    if not _LOGO_PATH.exists():
        return

    logo = XLImage(str(_LOGO_PATH))
    logo.width = 150
    logo.height = 70
    logo.anchor = "A1"
    ws.add_image(logo)


# ── Helpers ──────────────────────────────────────────────────────────────────

_MOIS_FR = [
    "", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
]

_FONCTION_LABEL = {
    "chef_atelier":       "SAID",
    "peintre":            "PEINTRE",
    "tolier":             "TOLIER",
    "mecanicien":         "MECANICIEN",
    "electricien":        "ELECTRICIEN",
    "diagnostic":         "DIAGNOSTIC",
    "ouvrier":            "OUVRIER",
    "administratif":      "ADMINIST",
    "mecanicien_nautique": "TACHE",
    "gerant":             "GERANT",
    "autre":              "AUTRE",
}

_TYPE_OBS = {
    "avance":   "AVANCE",
    "prime":    "PRIME",
    "credit":   "CREDIT",
    "cumul":    "CUMUL",
    "frais":    None,       # on garde la description telle quelle
    "reste_du": None,
    "tache":    None,
}


def _obs_label(avance: AvanceSalaire) -> str:
    """Observation à afficher dans la colonne E."""
    if avance.description:
        return avance.description.upper()
    mapped = _TYPE_OBS.get(avance.type)
    return mapped or avance.type.upper()


# ── Feuille SALAIRES ─────────────────────────────────────────────────────────

def _build_salaires(ws, mois: int, annee: int):
    dernier_jour = calendar.monthrange(annee, mois)[1]
    periode_debut = date(annee, mois, 1)
    periode_fin   = date(annee, mois, dernier_jour)

    # Largeurs des colonnes A–K
    widths = [22, 14, 14, 18, 38, 4, 4, 16, 4, 4, 4]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── En-tête atelier ─────────────────────────────────────────────────────
    for row_index, height in [(1, 24), (2, 26), (3, 22), (4, 22), (5, 8), (6, 10)]:
        ws.row_dimensions[row_index].height = height

    ws.merge_cells("A1:B4")
    _style_range(ws, "A1:B4", fill=_fill(_SLATE_900), border=_border_thin(), alignment=_align("center"))
    _add_logo(ws)

    ws.merge_cells("C1:K2")
    _style_range(ws, "C1:K2", fill=_fill(_SLATE_900), border=_border_thin(), alignment=_align("center"))
    ws["C1"].value = "MONSTER GARAGE"
    ws["C1"].font = _font(bold=True, size=24, color=_JAUNE)
    ws["C1"].alignment = _align("center")

    ws.merge_cells("C3:K3")
    _style_range(ws, "C3:K3", fill=_fill(_SLATE_900), border=_border_thin(), alignment=_align("center"))
    ws["C3"].value = "WIDINE MOTORS SERVICE"
    ws["C3"].font = _font(bold=True, size=12, color=_BLANC)
    ws["C3"].alignment = _align("center")

    ws.merge_cells("C4:K4")
    _style_range(ws, "C4:K4", fill=_fill(_YELLOW_SOFT), border=_border_thin(), alignment=_align("center"))
    ws["C4"].value = f"LIVRE DE PAIE — {_MOIS_FR[mois].upper()} {annee}    |    PERIODE DU {periode_debut:%d/%m/%Y} AU {periode_fin:%d/%m/%Y}"
    ws["C4"].font = _font(bold=True, size=11, color=_NOIR)
    ws["C4"].alignment = _align("center")

    # ── Ligne 7 : en-têtes colonnes ─────────────────────────────────────────
    headers = ["Nom & Prénom", "Date", "Montant", "Fonction", "Observations"]
    ws.row_dimensions[7].height = 22
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font      = _font(bold=True, size=10, color=_BLANC)
        cell.fill      = _fill(_SLATE_900)
        cell.alignment = _align("center")
        cell.border    = _border_thin()

    # ── Récupération des données ─────────────────────────────────────────────
    avances = (
        AvanceSalaire.query
        .join(Employe)
        .filter(AvanceSalaire.mois == mois, AvanceSalaire.annee == annee)
        .order_by(AvanceSalaire.date, Employe.nom_complet)
        .all()
    )

    salaires = (
        Salaire.query
        .join(Employe)
        .filter(Salaire.mois == mois, Salaire.annee == annee)
        .order_by(Salaire.date, Employe.nom_complet)
        .all()
    )

    quinzaines = [s for s in salaires if s.type_paie == "quinzaine"]
    fins_mois  = [s for s in salaires if s.type_paie == "fin_mois"]

    avances_q1 = [a for a in avances if a.quinzaine == "premiere"]
    avances_q2 = [a for a in avances if a.quinzaine == "seconde" or a.quinzaine is None]

    row = 8

    def write_avance(r, av, fill_hex=None):
        ws.row_dimensions[r].height = 15
        fn_label = _FONCTION_LABEL.get(av.employe.fonction, av.employe.fonction.upper())
        vals = [
            av.employe.nom_complet,
            av.date,
            float(av.montant),
            fn_label,
            _obs_label(av),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font      = _font(size=10)
            cell.alignment = _align("left" if col != 3 else "right")
            cell.border    = _border_bottom()
            if col == 2:
                cell.number_format = "DD/MM/YYYY"
            if col == 3:
                cell.number_format = '#,##0.00'
            if fill_hex:
                cell.fill = _fill(fill_hex)

    def write_salaire(r, sal, fill_hex=None):
        ws.row_dimensions[r].height = 15
        fn_label = _FONCTION_LABEL.get(sal.employe.fonction, sal.employe.fonction.upper())
        obs = sal.date   # colonne E = date (comme dans l'original)
        vals = [
            sal.employe.nom_complet,
            sal.date,
            float(sal.montant_net_paye),
            fn_label,
            obs,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font      = _font(bold=(col == 3), size=10)
            cell.alignment = _align("left" if col != 3 else "right")
            cell.border    = _border_bottom()
            if col in (2, 5):
                cell.number_format = "DD/MM/YYYY"
            if col == 3:
                cell.number_format = '#,##0.00'
            if fill_hex:
                cell.fill = _fill(fill_hex)

    def write_quinzaine_label(r, sal):
        ws.row_dimensions[r].height = 15
        fn_label = _FONCTION_LABEL.get(sal.employe.fonction, sal.employe.fonction.upper())
        obs = f"QUINZ {sal.date.strftime('%d/%m/%Y')}"
        vals = [sal.employe.nom_complet, sal.date, float(sal.montant_net_paye), fn_label, obs]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font      = _font(bold=True, size=10, color=_NOIR)
            cell.fill      = _fill(_JAUNE)
            cell.alignment = _align("left" if col != 3 else "right")
            cell.border    = _border_thin()
            if col == 2:
                cell.number_format = "DD/MM/YYYY"
            if col == 3:
                cell.number_format = '#,##0.00'

    # ── Q1 : avances et tâches avant le 15 ──────────────────────────────────
    for av in avances_q1:
        write_avance(row, av)
        row += 1

    # ── Quinzaines (15 du mois) ──────────────────────────────────────────────
    for sal in quinzaines:
        write_quinzaine_label(row, sal)
        row += 1

    # ── Séparateur total quinzaine ────────────────────────────────────────────
    total_q1 = sum(float(s.montant_net_paye) for s in quinzaines)
    ws.row_dimensions[row].height = 16
    ws.cell(row=row, column=1, value=date(annee, mois, 16)).number_format = "DD/MM/YYYY"
    ws.cell(row=row, column=3, value="Total                            :")
    total_cell = ws.cell(row=row, column=4, value=total_q1)
    total_cell.number_format = '#,##0.00'
    for col in range(1, 6):
        c = ws.cell(row=row, column=col)
        c.font = _font(bold=True, size=10)
        c.fill = _fill(_GRIS_LGT)
        c.border = _border_thin()
    row += 1

    # ── Q2 : avances après le 15 ─────────────────────────────────────────────
    for av in avances_q2:
        write_avance(row, av)
        row += 1

    # ── Soldes fin de mois ────────────────────────────────────────────────────
    for sal in fins_mois:
        write_salaire(row, sal, fill_hex=_GRIS_LGT)
        row += 1

    # ── Total général ─────────────────────────────────────────────────────────
    total_general = (
        sum(float(a.montant) for a in avances)
        + sum(float(s.montant_net_paye) for s in salaires)
    )
    row += 1
    ws.row_dimensions[row].height = 18
    ws.cell(row=row, column=3, value="TOTAL SALAIRES")
    total_g = ws.cell(row=row, column=5, value=total_general)
    total_g.number_format = '#,##0.00'
    for col in [3, 5]:
        c = ws.cell(row=row, column=col)
        c.font = _font(bold=True, size=11, color=_BLANC)
        c.fill = _fill(_GRIS_CLR)
        c.alignment = _align("center")
        c.border = _border_thin()

    # Ligne confirmation
    row += 2
    ws.cell(row=row, column=4, value="Total Salaires")
    conf = ws.cell(row=row, column=5, value=total_general)
    conf.number_format = '#,##0.00'
    conf.font = _font(bold=True, size=11)


# ── Point d'entrée principal ─────────────────────────────────────────────────

def exporter_salaires_excel(mois: int, annee: int) -> bytes:
    """Génère le fichier Excel SALAIRES pour le mois donné. Retourne les bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SALAIRES"
    ws.sheet_view.showGridLines = False

    _build_salaires(ws, mois, annee)

    # Feuilles vides pour garder la structure du fichier original
    for name in ("JOURNAL DE CAISSE", "DETAIL CA MOIS", "BCE", "REPART CA"):
        wb.create_sheet(name)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
