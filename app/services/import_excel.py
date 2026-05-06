"""
Import du livre de caisse/salaires depuis le fichier Excel mensuel.

Colonnes lues dans la feuille SALAIRES :
  A = Nom & Prénom   B = Date   C = Montant   D = Fonction   E = Observations
"""

import re
import calendar
from datetime import date, datetime
from decimal import Decimal

import openpyxl

from app.extensions import db
from app.models.avance_salaire import AvanceSalaire
from app.models.employe import Employe
from app.models.salaire import Salaire

# ---------------------------------------------------------------------------
# Mapping fonction Excel → enum SQLAlchemy
# ---------------------------------------------------------------------------

_FONCTION_MAP = {
    "SAID":              "chef_atelier",
    "CHEF ATELIER":      "chef_atelier",
    "PEINTRE":           "peintre",
    "TOLIER":            "tolier",
    "TÔLIER":            "tolier",
    "OUVRIER":           "ouvrier",
    "ADMINIST":          "administratif",
    "ADMINISTRATIF":     "administratif",
    "ELECTRICIEN":       "electricien",
    "ÉLECTRICIEN":       "electricien",
    "DIAGNOSTIC":        "diagnostic",
    "MECANICIEN":        "mecanicien",
    "MÉCANICIEN":        "mecanicien",
    "MECANICIEN NAUTIQUE": "mecanicien_nautique",
    "TACHE":             "ouvrier",
    "TÂCHE":             "ouvrier",
}

_SALAIRE_QUINZAINE_CONNU = {
    "SAID":       Decimal("3700"),
    "SIMOHAMED":  Decimal("2700"),
    "ABDESLAM":   Decimal("2250"),
    "ABDELLHADI": Decimal("2250"),
    "ZOUHEIR":    Decimal("1600"),
    "OUTHMANE":   Decimal("1600"),
    "SALIMA":     Decimal("1500"),
    "FENNI":      Decimal("1200"),
}


# ---------------------------------------------------------------------------
# Détection du type d'opération
# ---------------------------------------------------------------------------

def _detecter_type(obs) -> str:
    """Détermine le type d'opération depuis la colonne Observations."""
    if obs is None:
        return "tache"
    if isinstance(obs, (datetime, date)):
        return "fin_mois"
    obs_upper = str(obs).strip().upper()
    if obs_upper.startswith("QUINZ"):
        return "quinzaine"
    if "AVANCE" in obs_upper:
        return "avance"
    if "PRIME" in obs_upper:
        return "prime"
    if "CREDIT" in obs_upper or "CRÉDIT" in obs_upper:
        return "credit"
    if "CUMUL" in obs_upper:
        return "cumul"
    if any(k in obs_upper for k in ("DEJEUNER", "DÉJEUNER", "REPAS", "AMINE")):
        return "frais"
    return "tache"


def _extraire_reste_du(obs) -> Decimal | None:
    """Extrait le montant 'RESTE = XXXX' depuis les observations."""
    if not obs or not isinstance(obs, str):
        return None
    m = re.search(r"RESTE\s*=\s*([\d\s]+)", obs.upper())
    if m:
        try:
            return Decimal(m.group(1).replace(" ", ""))
        except Exception:
            return None
    return None


def _mapper_fonction(valeur: str | None) -> str:
    if not valeur:
        return "autre"
    return _FONCTION_MAP.get(str(valeur).strip().upper(), "autre")


# ---------------------------------------------------------------------------
# Résolution / création d'employé
# ---------------------------------------------------------------------------

def _trouver_ou_creer_employe(nom: str, fonction_excel: str,
                               cache: dict) -> tuple[Employe, bool]:
    """
    Cherche l'employé par nom (insensible à la casse + strip).
    Le crée s'il n'existe pas.
    Retourne (employe, cree).
    """
    nom_norm = nom.strip().upper()
    if nom_norm in cache:
        return cache[nom_norm], False

    emp = Employe.query.filter(
        db.func.upper(db.func.trim(Employe.nom_complet)) == nom_norm
    ).first()

    cree = False
    if not emp:
        fonction = _mapper_fonction(fonction_excel)
        salaire_q = _SALAIRE_QUINZAINE_CONNU.get(nom_norm)
        type_rem = "salaire_fixe" if salaire_q else "tache"
        if nom_norm == "SAID":
            type_rem = "mixte"

        emp = Employe(
            nom_complet=nom_norm,
            fonction=fonction,
            type_remuneration=type_rem,
            salaire_quinzaine=salaire_q,
            actif=True,
        )
        db.session.add(emp)
        db.session.flush()   # obtenir l'id
        cree = True

    cache[nom_norm] = emp
    return emp, cree


def _operation_deja_enregistree(
    employe_id: int,
    date_op: date,
    montant: Decimal,
    type_op: str,
    description: str | None,
    quinzaine: str,
    mois: int,
    annee: int,
) -> bool:
    """Evite de dupliquer les lignes d'operations lors d'un re-import Excel."""
    query = AvanceSalaire.query.filter(
        AvanceSalaire.employe_id == employe_id,
        AvanceSalaire.date == date_op,
        AvanceSalaire.montant == montant,
        AvanceSalaire.type == type_op,
        AvanceSalaire.quinzaine == quinzaine,
        AvanceSalaire.mois == mois,
        AvanceSalaire.annee == annee,
    )
    if description:
        query = query.filter(AvanceSalaire.description == description)
    else:
        query = query.filter(AvanceSalaire.description.is_(None))
    return query.first() is not None


# ---------------------------------------------------------------------------
# Fonction principale d'import
# ---------------------------------------------------------------------------

def importer_salaires_excel(
    source,           # chemin str ou bytes-like
    mois: int,
    annee: int,
    nom_feuille: str = "SALAIRES",
    mode_sec: bool = False,   # True = dry-run, ne commit pas
) -> dict:
    """
    Lit la feuille Excel et insère les opérations en base.

    Retourne un dict récapitulatif :
    {
        'employes_crees': [str],
        'avances_creees': int,
        'salaires_crees': int,
        'ignores': [str],
        'erreurs': [str],
        'total_importe': Decimal,
    }
    """
    wb = openpyxl.load_workbook(source, data_only=True)
    if nom_feuille not in wb.sheetnames:
        raise ValueError(f"Feuille '{nom_feuille}' introuvable. Feuilles disponibles : {wb.sheetnames}")

    ws = wb[nom_feuille]

    employes_cache: dict[str, Employe] = {}
    employes_crees: list[str] = []
    avances_creees = 0
    salaires_crees = 0
    ignores: list[str] = []
    erreurs: list[str] = []
    total_importe = Decimal("0")

    dernier_jour = calendar.monthrange(annee, mois)[1]

    for row in ws.iter_rows(min_row=1, values_only=True):
        row = tuple(row) + (None, None, None, None, None)
        nom_raw, date_raw, montant_raw, fonction_raw, obs_raw = (
            row[0], row[1], row[2], row[3], row[4]
        )

        # ── Ignorer les lignes sans nom ou sans montant valide ──────────
        if not nom_raw or not isinstance(nom_raw, str):
            continue
        nom = nom_raw.strip().upper()
        if not nom or nom in ("NOM & PRÉNOM", "NOM & PR�NOM", "MONSTER GARAGE",
                               "LIVRE DE CAISSE", "RECAP SALAIRES"):
            continue
        if montant_raw is None or not isinstance(montant_raw, (int, float)):
            continue
        montant = Decimal(str(montant_raw))
        if montant <= 0:
            continue

        # ── Date ────────────────────────────────────────────────────────
        if isinstance(date_raw, datetime):
            date_op = date_raw.date()
        elif isinstance(date_raw, date):
            date_op = date_raw
        else:
            date_op = date(annee, mois, dernier_jour)   # fallback pour fin de mois

        # Vérification cohérence mois/année
        if date_op.month != mois or date_op.year != annee:
            ignores.append(f"{nom} {date_op} — hors période {mois}/{annee}")
            continue

        type_op = _detecter_type(obs_raw)

        try:
            emp, cree = _trouver_ou_creer_employe(nom, str(fonction_raw or ""), employes_cache)
            if cree:
                employes_crees.append(nom)

            quinzaine = "premiere" if date_op.day <= 15 else "seconde"

            # ── Salaire : quinzaine ──────────────────────────────────────
            if type_op == "quinzaine":
                existe = Salaire.query.filter_by(
                    employe_id=emp.id, mois=mois, annee=annee, type_paie="quinzaine"
                ).first()
                if existe:
                    ignores.append(f"{nom} — quinzaine déjà enregistrée")
                    continue

                # La ligne QUINZ est le montant de référence par quinzaine.
                if emp.salaire_quinzaine is None:
                    emp.salaire_quinzaine = montant

                sal = Salaire(
                    employe_id=emp.id,
                    type_paie="quinzaine",
                    date=date_op,
                    mois=mois,
                    annee=annee,
                    salaire_brut=montant,
                    total_avances=Decimal("0"),
                    total_primes=Decimal("0"),
                    montant_net_paye=montant,
                )
                db.session.add(sal)
                salaires_crees += 1
                total_importe += montant

            # ── Salaire : fin de mois ────────────────────────────────────
            elif type_op == "fin_mois":
                existe = Salaire.query.filter_by(
                    employe_id=emp.id, mois=mois, annee=annee, type_paie="fin_mois"
                ).first()
                if existe:
                    ignores.append(f"{nom} — solde fin de mois déjà enregistré")
                    continue

                sal = Salaire(
                    employe_id=emp.id,
                    type_paie="fin_mois",
                    date=date_op,
                    mois=mois,
                    annee=annee,
                    salaire_brut=montant,
                    total_avances=Decimal("0"),
                    total_primes=Decimal("0"),
                    montant_net_paye=montant,
                )
                db.session.add(sal)
                salaires_crees += 1
                total_importe += montant

            # ── Avance / prime / tâche / frais / cumul ───────────────────
            else:
                description = str(obs_raw).strip() if obs_raw and not isinstance(obs_raw, (datetime, date)) else ""
                description_db = description or None

                if _operation_deja_enregistree(
                    emp.id,
                    date_op,
                    montant,
                    type_op,
                    description_db,
                    quinzaine,
                    mois,
                    annee,
                ):
                    ignores.append(f"{nom} — opération déjà enregistrée ({type_op}, {date_op})")
                    continue

                reste_du = None
                montant_total_convenu = None
                if type_op == "tache":
                    reste_du = _extraire_reste_du(description)
                    if reste_du:
                        montant_total_convenu = montant + reste_du

                av = AvanceSalaire(
                    employe_id=emp.id,
                    date=date_op,
                    montant=montant,
                    type=type_op,
                    description=description_db,
                    quinzaine=quinzaine,
                    mois=mois,
                    annee=annee,
                    montant_total_convenu=montant_total_convenu,
                    reste_du=reste_du,
                )
                db.session.add(av)
                avances_creees += 1
                total_importe += montant

        except Exception as exc:
            erreurs.append(f"{nom} ligne {date_raw} : {exc}")

    if mode_sec:
        db.session.rollback()
    else:
        db.session.commit()

    return {
        "employes_crees": employes_crees,
        "avances_creees": avances_creees,
        "salaires_crees": salaires_crees,
        "ignores": ignores,
        "erreurs": erreurs,
        "total_importe": total_importe,
    }
