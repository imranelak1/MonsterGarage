from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import Client, DevisReparation, FactureReparation
from app.services.calculs_facturation import calculer_commission_sntl, param_percent
from app.services.devis_totaux import calculer_totaux_lignes, montant_ttc_ligne, taux_tva_ligne
from app.services.parametres import obtenir_entreprise


XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_JAUNE = "FFFFD700"
_NOIR = "FF000000"
_BLANC = "FFFFFFFF"
_SLATE_950 = "FF020617"
_SLATE_100 = "FFF1F5F9"
_SLATE_200 = "FFE2E8F0"
_SLATE_500 = "FF64748B"
_AMBRE_PALE = "FFFFF7CC"
_VERT_PALE = "FFEAF7EE"
_ROUGE_PALE = "FFFFECEC"
_LOGO_PATH = Path(__file__).resolve().parent.parent / "static" / "img" / "logo_monster_garage.png"
_SNTL_FONT = "Aptos Narrow"


def exporter_devis_excel(devis: DevisReparation) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"DEVIS V{devis.version}"
    is_sntl = devis.dossier.client.type == "sntl"
    _document_setup_devis_modele(ws, is_sntl=is_sntl)
    _build_devis_modele(ws, devis, is_sntl=is_sntl)
    return _save(wb)


def exporter_facture_excel(facture: FactureReparation) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(facture.numero)

    if facture.dossier.client.type == "sntl":
        _document_setup_sntl(ws)
        _build_facture_sntl(ws, facture)
    else:
        _document_setup(ws)
        _build_document_reparation(
            ws,
            titre="FACTURE",
            numero=facture.numero,
            date_document=facture.created_at,
            devis=facture.devis,
            statut=facture.statut_libelle,
            lignes=facture.lignes_facture,
            montant_ht=facture.montant_ht,
            montant_tva=facture.montant_tva,
            montant_ttc=facture.montant_ttc,
            montant_regle=facture.montant_regle,
            hide_etat=False,
            include_commission_sntl=False,
        )
    return _save(wb)


def exporter_releve_client_excel(client: Client, factures: list[FactureReparation]) -> bytes:
    wb = Workbook()
    ws = wb.active
    if client.type == "sntl":
        ws.title = "RELEVÉ SNTL"
        _build_releve_sntl(ws, client, factures)
        return _save(wb)

    _build_releve_client(ws, client, factures)
    _build_resume_vehicules(wb, client, factures)
    return _save(wb)


def exporter_situation_clients_excel(clients: list[Client], factures: list[FactureReparation]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "SITUATION GLOBALE"
    _document_setup(ws, widths=[30, 18, 14, 18, 18, 18, 18, 18, 18])
    _build_header(ws, "SITUATION FINANCIERE CLIENTS", "GLOBAL", date.today())

    headers = [
        "CLIENT",
        "TYPE",
        "NB FACTURES",
        "MONTANT TTC",
        "COMMISSION SNTL",
        "NET À RÉGLER",
        "ENCAISSÉ",
        "À ENCAISSER",
        "DERNIÈRE FACTURE",
    ]
    header_row = 9
    _write_table_header(ws, header_row, headers)

    factures_par_client = {client.id: [] for client in clients}
    for facture in factures:
        factures_par_client.setdefault(facture.dossier.client_id, []).append(facture)

    first_data_row = header_row + 1
    row = first_data_row
    totals = {
        3: 0,
        4: Decimal("0.00"),
        5: Decimal("0.00"),
        6: Decimal("0.00"),
        7: Decimal("0.00"),
        8: Decimal("0.00"),
    }
    for client in clients:
        client_factures = factures_par_client.get(client.id, [])
        if not client_factures:
            continue

        total_ttc = sum((Decimal(str(facture.montant_ttc or 0)) for facture in client_factures), Decimal("0.00"))
        commission = sum((_commission_releve_sntl(facture) for facture in client_factures), Decimal("0.00"))
        net_a_regler = total_ttc - commission
        encaisse = sum((Decimal(str(facture.montant_regle or 0)) for facture in client_factures), Decimal("0.00"))
        reste = max(net_a_regler - encaisse, Decimal("0.00"))
        derniere_facture = max(client_factures, key=lambda facture: facture.created_at)
        totals[3] += len(client_factures)
        totals[4] += total_ttc
        totals[5] += commission
        totals[6] += net_a_regler
        totals[7] += encaisse
        totals[8] += reste

        values = [
            client.nom,
            client.type_libelle,
            len(client_factures),
            _money(total_ttc),
            _money(commission),
            _money(net_a_regler),
            _money(encaisse),
            _money(reste),
            _date_value(derniere_facture.created_at),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.border = _border("FFCBD5E1")
            cell.alignment = _align("center" if col != 1 else "left", wrap=True)
            if col in {4, 5, 6, 7, 8}:
                cell.number_format = '#,##0.00 "MAD"'
            if col == 9:
                cell.number_format = "dd/mm/yyyy"

        detail_title = f"SNTL - {client.nom}" if client.type == "sntl" else f"SITUATION - {client.nom}"
        detail_ws = wb.create_sheet(_unique_sheet_title(wb, detail_title))
        if client.type == "sntl":
            _build_releve_sntl(detail_ws, client, client_factures)
        else:
            _build_releve_client(detail_ws, client, client_factures)

        row += 1

    total_row = max(first_data_row, row) + 1
    ws.cell(total_row, 2, "TOTAL").font = _font(bold=True, size=12)
    for col in (3, 4, 5, 6, 7, 8):
        value = totals[col] if col == 3 else _money(totals[col])
        cell = ws.cell(total_row, col, value)
        cell.font = _font(bold=True, size=12)
        cell.fill = _fill(_AMBRE_PALE)
        cell.border = _border()
        if col in {4, 5, 6, 7, 8}:
            cell.number_format = '#,##0.00 "MAD"'

    _set_print_options(ws)
    return _save(wb)


def _build_releve_client(ws, client: Client, factures: list[FactureReparation]) -> None:
    if ws.title == "Sheet":
        ws.title = "SITUATION CLIENT"
    _document_setup(ws, widths=[18, 28, 24, 18, 18, 18, 18, 18])
    _build_header(ws, f"SITUATION {client.nom.upper()}", client.type_libelle.upper(), date.today())

    ws.merge_cells("A6:D6")
    ws["A6"].value = "WIDINE MOTORS SERVICES"
    ws["A6"].font = _font(bold=True, size=12, color=_SLATE_950)
    ws["A6"].alignment = _align("center")

    headers = ["N° FACTURE", "VÉHICULE", "IMMATRICULATION", "MONTANT FACTURE", "ENCAISSÉ", "À ENCAISSER", "DATE", "STATUT"]
    header_row = 9
    _write_table_header(ws, header_row, headers)

    first_data_row = header_row + 1
    total_montant = Decimal("0.00")
    total_encaisse = Decimal("0.00")
    for index, facture in enumerate(factures, first_data_row):
        vehicule = facture.dossier.vehicule
        montant_ttc = Decimal(str(facture.montant_ttc or 0)).quantize(Decimal("0.01"))
        encaisse = Decimal(str(facture.montant_regle or 0)).quantize(Decimal("0.01"))
        total_montant += montant_ttc
        total_encaisse += encaisse
        values = [
            facture.numero,
            f"{vehicule.marque} {vehicule.modele}",
            vehicule.immatriculation,
            _money(montant_ttc),
            _money(encaisse),
            f"=D{index}-E{index}",
            _date_value(facture.created_at),
            facture.statut_libelle,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(index, col, value)
            cell.border = _border("FFCBD5E1")
            cell.alignment = _align("center" if col != 2 else "left", wrap=True)
            if col in {4, 5, 6}:
                cell.number_format = '#,##0.00 "MAD"'

    total_row = max(first_data_row, first_data_row + len(factures)) + 1
    ws.cell(total_row, 3, "TOTAL").font = _font(bold=True, size=12)
    total_values = {
        4: total_montant,
        5: total_encaisse,
        6: total_montant - total_encaisse,
    }
    for col in (4, 5, 6):
        cell = ws.cell(total_row, col, _money(total_values[col]))
        cell.font = _font(bold=True, size=12)
        cell.fill = _fill(_AMBRE_PALE)
        cell.border = _border()
        cell.number_format = '#,##0.00 "MAD"'

    _set_print_options(ws)


def _build_releve_sntl(ws, client: Client, factures: list[FactureReparation]) -> None:
    _document_setup_sntl(ws)
    _build_sntl_releve_header(ws)

    headers = ["N° Facture", "Date facture", "N° Bon SNTL", "Montant TTC", "Commission SNTL", "Montant à régler"]
    header_row = 31
    for col, value in enumerate(headers, 2):
        cell = ws.cell(header_row, col, value)
        cell.font = _font(name=_SNTL_FONT, size=12)
        cell.alignment = _align("center", wrap=True)
        cell.border = _border()

    first_data_row = header_row + 1
    total_ttc = Decimal("0.00")
    total_commission = Decimal("0.00")
    total_net_a_regler = Decimal("0.00")
    for row, facture in enumerate(factures, first_data_row):
        commission = _commission_releve_sntl(facture)
        montant_ttc = Decimal(str(facture.montant_ttc or 0)).quantize(Decimal("0.01"))
        net_a_regler = montant_ttc - commission
        total_ttc += montant_ttc
        total_commission += commission
        total_net_a_regler += net_a_regler
        values = [
            facture.numero,
            _date_value(facture.created_at),
            _sntl_bon_number(facture.dossier, client),
            _money(montant_ttc),
            _money(commission),
            _money(net_a_regler),
        ]
        for col, value in enumerate(values, 2):
            cell = ws.cell(row, col, value)
            cell.border = _border()
            cell.alignment = _align("center", wrap=True)
            cell.font = _font(name=_SNTL_FONT, size=12)
            if col == 3:
                cell.number_format = "dd/mm/yyyy"
            if col in {5, 6, 7}:
                cell.number_format = '#,##0.00'

    total_row = max(first_data_row, first_data_row + len(factures)) + 1
    ws.cell(total_row, 2, "Total")
    ws.cell(total_row, 2).font = _font(name=_SNTL_FONT, bold=True, size=12)
    total_values = {
        5: total_ttc,
        6: total_commission,
        7: total_net_a_regler,
    }
    for col in (5, 6, 7):
        cell = ws.cell(total_row, col, _money(total_values[col]))
        cell.font = _font(name=_SNTL_FONT, bold=True, size=12)
        cell.border = _border()
        cell.number_format = '#,##0.00'
        cell.alignment = _align("right")

    signature_row = total_row + 3
    ws.cell(signature_row, 5, "Cachet et Signature")
    ws.cell(signature_row, 5).font = _font(name=_SNTL_FONT, size=12)
    ws.cell(signature_row, 5).alignment = _align()
    ws.print_area = f"A1:G{signature_row}"


def nom_fichier_devis(devis: DevisReparation) -> str:
    return _safe_filename(f"DEVIS_{devis.dossier.numero}_V{devis.version}.xlsx")


def nom_fichier_facture(facture: FactureReparation) -> str:
    prefix = "FACTURE_SNTL" if facture.dossier.client.type == "sntl" else "FACTURE"
    return _safe_filename(f"{prefix}_{facture.numero}.xlsx")


def nom_fichier_releve_client(client: Client) -> str:
    return _safe_filename(f"RELEVE_SITUATION_{client.nom}.xlsx")


def _build_resume_vehicules(wb: Workbook, client: Client, factures: list[FactureReparation]) -> None:
    ws = wb.create_sheet("RÉSUMÉ VÉHICULES")
    _document_setup(ws, widths=[28, 22, 16, 18, 18, 18, 18])
    _build_header(ws, f"RÉSUMÉ VÉHICULES - {client.nom.upper()}", client.type_libelle.upper(), date.today())

    headers = ["VÉHICULE", "IMMATRICULATION", "NB FACTURES", "MONTANT FACTURE", "ENCAISSÉ", "À ENCAISSER", "DERNIÈRE FACTURE"]
    header_row = 7
    for col, value in enumerate(headers, 1):
        cell = ws.cell(header_row, col, value)
        cell.fill = _fill(_SLATE_950)
        cell.font = _font(bold=True, color=_BLANC)
        cell.alignment = _align("center", wrap=True)
        cell.border = _border()

    grouped = {}
    for facture in factures:
        vehicule = facture.dossier.vehicule
        key = vehicule.id
        current = grouped.setdefault(
            key,
            {
                "vehicule": f"{vehicule.marque} {vehicule.modele}",
                "immatriculation": vehicule.immatriculation,
                "count": 0,
                "montant": Decimal("0.00"),
                "encaisse": Decimal("0.00"),
                "last": None,
            },
        )
        current["count"] += 1
        current["montant"] += Decimal(str(facture.montant_ttc or 0))
        current["encaisse"] += Decimal(str(facture.montant_regle or 0))
        last_date = _date_value(facture.created_at)
        if current["last"] is None or last_date > current["last"]:
            current["last"] = last_date

    first_row = header_row + 1
    for row, item in enumerate(grouped.values(), first_row):
        reste = item["montant"] - item["encaisse"]
        values = [
            item["vehicule"],
            item["immatriculation"],
            item["count"],
            _money(item["montant"]),
            _money(item["encaisse"]),
            _money(reste),
            item["last"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.border = _border("FFCBD5E1")
            cell.alignment = _align("center" if col != 1 else "left", wrap=True)
            if col in {4, 5, 6}:
                cell.number_format = '#,##0.00 "MAD"'
            if col == 7:
                cell.number_format = "dd/mm/yyyy"

    total_row = max(first_row, first_row + len(grouped)) + 1
    ws.cell(total_row, 2, "TOTAL").font = _font(bold=True, size=12)
    totals = {
        3: 0,
        4: Decimal("0.00"),
        5: Decimal("0.00"),
        6: Decimal("0.00"),
    }
    for item in grouped.values():
        totals[3] += item["count"]
        totals[4] += item["montant"]
        totals[5] += item["encaisse"]
        totals[6] += item["montant"] - item["encaisse"]
    for col in (3, 4, 5, 6):
        value = totals[col] if col == 3 else _money(totals[col])
        cell = ws.cell(total_row, col, value)
        cell.font = _font(bold=True, size=12)
        cell.fill = _fill(_AMBRE_PALE)
        cell.border = _border()
        if col in {4, 5, 6}:
            cell.number_format = '#,##0.00 "MAD"'

    _set_print_options(ws)


def _build_devis_sntl(ws, devis: DevisReparation) -> None:
    _build_sntl_official_document(
        ws,
        devis=devis,
        titre="DEVIS SNTL",
        numero=f"{devis.dossier.numero}-V{devis.version}",
        date_document=devis.created_at,
        include_commission=False,
    )


def _build_devis_modele(ws, devis: DevisReparation, *, is_sntl: bool) -> None:
    dossier = devis.dossier
    client = dossier.client
    vehicule = dossier.vehicule
    entreprise = obtenir_entreprise()
    lignes = list(devis.lignes)
    montant_ht, montant_tva, montant_ttc = calculer_totaux_lignes(lignes, client.type)
    max_col = 7

    ws.merge_cells("A1:D3")
    ws["A1"].value = "MONSTER GARAGE"
    ws["A1"].font = _font(bold=True, italic=True, size=24, color="FF2A93B0")
    ws["A1"].alignment = _align("center")
    _add_logo_at(ws, "F1", width=130, height=61)
    for col in range(1, 8):
        ws.cell(4, col).border = Border(bottom=Side(style="medium", color=_NOIR))

    if is_sntl:
        _devis_info_box(
            ws,
            6,
            1,
            [
                ("Devis n°", _numero_devis_modele(devis)),
                ("Date", _date_value(devis.created_at)),
                ("Véhicule", f"{vehicule.marque} {vehicule.modele}".upper()),
                ("Immatriculation", vehicule.immatriculation),
                ("Kilométrage", dossier.kilometrage_entree or vehicule.kilometrage_actuel or ""),
            ],
            width=3,
        )
        _devis_info_box(
            ws,
            6,
            5,
            [
                ("N°client", client.code),
                ("Client", (client.administration_rattachee or client.nom).upper()),
                ("Adresse", client.adresse or client.ville or ""),
            ],
            width=3,
        )
        header_row = 13
        headers = ["Désignation", "Qté", "Prix HT", "Prix HT total", "TVA", "Total TTC"]
    else:
        _devis_info_lines(
            ws,
            6,
            1,
            [
                ("Devis n°", _numero_devis_modele(devis)),
                ("Date", _date_value(devis.created_at)),
                ("Véhicule", f"{vehicule.marque} {vehicule.modele}".upper()),
                ("Immatriculation", vehicule.immatriculation),
            ],
        )
        _devis_info_lines(
            ws,
            6,
            5,
            [
                ("Client", client.nom.upper()),
                ("ICE n°", client.ice or ""),
                ("Adresse", client.adresse or client.ville or ""),
            ],
        )
        header_row = 12
        headers = ["Désignation", "Qté", "Etat", "Prix HT", "Prix HT total", "TVA", "Total TTC"]

    for offset, label in enumerate(headers, 1):
        cell = ws.cell(header_row, offset, label)
        cell.font = _font(bold=True, size=9)
        cell.alignment = _align("center", wrap=True)
        cell.border = _border()

    first_line_row = header_row + 1
    rows_count = max(10, len(lignes))
    for offset in range(rows_count):
        row = first_line_row + offset
        ligne = lignes[offset] if offset < len(lignes) else None
        values = _devis_modele_row_values(ligne, client.type, show_etat=not is_sntl)
        ws.row_dimensions[row].height = 16
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.font = _font(size=8)
            cell.alignment = _align("center", wrap=True)
            cell.border = _border()
            if col in ({3, 4, 6} if is_sntl else {4, 5, 7}):
                cell.number_format = '#,##0.00'

    totals_row = first_line_row + rows_count + 1
    label_col = len(headers) - 1
    value_col = len(headers)
    for offset, (label, value) in enumerate((("Total HT", montant_ht), ("TVA", montant_tva), ("Total TTC", montant_ttc))):
        row = totals_row + offset
        ws.cell(row, label_col, label)
        ws.cell(row, value_col, float(value))
        ws.cell(row, label_col).font = _font(bold=True, size=8)
        ws.cell(row, value_col).font = _font(size=8)
        ws.cell(row, value_col).number_format = '#,##0.00"MAD"'
        _style_range(ws, row, label_col, row, value_col, border=_border(), alignment=_align("center"))

    legal_row = totals_row + 6
    ws.merge_cells(start_row=legal_row, start_column=1, end_row=legal_row, end_column=max_col)
    ws.cell(legal_row, 1, f"Arrêté le présent devis à la somme de: {_capitalize(_amount_to_french(montant_ttc))} TTC")
    ws.cell(legal_row, 1).font = _font(bold=True, italic=True, size=8)
    ws.cell(legal_row, 1).alignment = _align("left", wrap=True)
    ws.cell(legal_row + 2, 1, "Devis reçu avant l'exécution des travaux")
    ws.cell(legal_row + 2, 1).font = _font(italic=True, size=8)

    signature_row = legal_row + 8
    ws.cell(signature_row, 5, "Date :")
    ws.cell(signature_row + 2, 5, "Bon pour accord")
    ws.cell(signature_row + 4, 5, "Signature client")
    for row in (signature_row, signature_row + 2, signature_row + 4):
        ws.cell(row, 5).font = _font(size=8)
        ws.cell(row, 5).alignment = _align("center")

    footer_row = signature_row + 9
    footer_1 = _join_parts(f"RC: {entreprise.rc}", f"IF: {entreprise.if_fiscal}", f"ICE: {entreprise.ice}", f"Patente: {entreprise.patente}", sep=" / ")
    footer_2 = f"Adresse: {entreprise.adresse}, {entreprise.ville}"
    footer_3 = _join_parts(f"Tél: {entreprise.telephones}", f"E-mail: {entreprise.email}", sep=" / ")
    for offset, text in enumerate((footer_1, footer_2, footer_3)):
        row = footer_row + offset
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        ws.cell(row, 1, text)
        ws.cell(row, 1).font = _font(size=7)
        ws.cell(row, 1).alignment = _align("center")
    for col in range(1, max_col + 1):
        ws.cell(footer_row - 1, col).border = Border(bottom=Side(style="medium", color=_NOIR))

    ws.print_area = f"A1:{get_column_letter(max_col)}{footer_row + 2}"
    _set_print_options(ws)


def _devis_info_box(ws, row: int, col: int, rows: list[tuple[str, object]], *, width: int) -> None:
    end_col = col + width - 1
    end_row = row + len(rows) - 1
    _style_range(ws, row, col, end_row, end_col, border=_border("FF808080"))
    for index, (label, value) in enumerate(rows, row):
        ws.cell(index, col, f"{label} :")
        ws.cell(index, col + 1, _display(value))
        if isinstance(value, date):
            ws.cell(index, col + 1).number_format = "dd/mm/yyyy"
        ws.cell(index, col).font = _font(size=8, color=_SLATE_500)
        ws.cell(index, col + 1).font = _font(bold=True, size=8)
        ws.cell(index, col).alignment = _align("left")
        ws.cell(index, col + 1).alignment = _align("left", wrap=True)
        if width > 2:
            ws.merge_cells(start_row=index, start_column=col + 1, end_row=index, end_column=end_col)


def _devis_info_lines(ws, row: int, col: int, rows: list[tuple[str, object]]) -> None:
    for index, (label, value) in enumerate(rows, row):
        ws.cell(index, col, f"{label}:")
        ws.cell(index, col + 1, _display(value))
        if isinstance(value, date):
            ws.cell(index, col + 1).number_format = "dd/mm/yyyy"
        ws.cell(index, col).font = _font(size=8, color=_SLATE_500)
        ws.cell(index, col + 1).font = _font(bold=True, size=8)
        ws.cell(index, col).alignment = _align("left")
        ws.cell(index, col + 1).alignment = _align("left", wrap=True)


def _devis_modele_row_values(ligne, client_type: str, *, show_etat: bool) -> list:
    if ligne is None:
        return [""] * (7 if show_etat else 6)

    values = [
        str(ligne.designation or "").upper(),
        float(Decimal(str(ligne.quantite or 0))),
        float(Decimal(str(ligne.prix_unitaire_ht or 0))),
        float(Decimal(str(ligne.total_ht or 0))),
        _percent_label(taux_tva_ligne(ligne, client_type)),
        float(montant_ttc_ligne(ligne, client_type)),
    ]
    if show_etat:
        values.insert(2, _etat_modele(ligne))
    return values


def _document_setup_devis_modele(ws, *, is_sntl: bool) -> None:
    widths = [26, 8, 11, 14, 16, 9, 16] if not is_sntl else [30, 8, 14, 16, 9, 16, 2]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in range(1, 55):
        ws.row_dimensions[row].height = 15
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35


def _add_logo_at(ws, anchor: str, *, width: int, height: int) -> None:
    if not _LOGO_PATH.exists():
        return
    logo = XLImage(str(_LOGO_PATH))
    logo.width = width
    logo.height = height
    logo.anchor = anchor
    ws.add_image(logo)


def _numero_devis_modele(devis: DevisReparation) -> str:
    return f"{devis.dossier.numero}-V{devis.version}"


def _etat_modele(ligne) -> str:
    if getattr(ligne, "type_ligne", None) == "main_oeuvre" or getattr(ligne, "etat_piece", None) == "mo":
        return "MO"
    if getattr(ligne, "etat_piece", None) == "occasion":
        return "OCC"
    if getattr(ligne, "etat_piece", None) == "autre":
        return str(getattr(ligne, "etat_piece_autre", "") or "AUTRE").upper()
    return "NEUF"


def _percent_label(rate: Decimal) -> str:
    return f"{int((Decimal(str(rate)) * 100).quantize(Decimal('1')))}%"


def _build_facture_sntl(ws, facture: FactureReparation) -> None:
    _build_sntl_official_document(
        ws,
        devis=facture.devis,
        titre="FACTURE SNTL",
        numero=facture.numero,
        date_document=facture.created_at,
        include_commission=True,
        lignes=facture.lignes_facture,
    )


def _build_sntl_official_document(
    ws,
    *,
    devis: DevisReparation,
    titre: str,
    numero: str,
    date_document,
    include_commission: bool,
    lignes=None,
) -> None:
    dossier = devis.dossier
    client = dossier.client
    vehicule = dossier.vehicule
    entreprise = obtenir_entreprise()
    tva = _param_percent("taux_tva", Decimal("20"))
    commission = _param_percent("taux_commission_sntl", Decimal("10"))

    _build_sntl_official_header(ws, titre, numero, date_document)

    _sntl_section_header(ws, "B18:C18", "Partenaire")
    _sntl_section_header(ws, "D18:E18", "Véhicule")
    _sntl_section_header(ws, "F18:G18", "Partenaire")

    partenaire_rows = [
        ("N° Agrément SNTL", entreprise.agrement_sntl),
        ("Raison Sociale", entreprise.raison_sociale),
        ("Adresse", entreprise.adresse),
        ("Ville", entreprise.ville),
        ("RC", entreprise.rc),
        ("Patente", entreprise.patente),
        ("IF", entreprise.if_fiscal),
        ("ICE", entreprise.ice),
        ("N° RIB", entreprise.rib),
    ]
    for row, (label, value) in enumerate(partenaire_rows, 19):
        _sntl_label_value(ws, row, 2, label, value)

    vehicle_rows = [
        ("Matricule", _format_sntl_matricule(vehicule.immatriculation)),
        ("Marque et modèle", f"{vehicule.marque} {vehicule.modele}".upper()),
        ("Kilométrage", dossier.kilometrage_entree or vehicule.kilometrage_actuel),
        ("Administration", (client.administration_rattachee or client.nom).upper()),
        ("N° de bon", _sntl_bon_number(dossier, client)),
    ]
    for row, (label, value) in enumerate(vehicle_rows, 19):
        _sntl_label_value(ws, row, 4, label, value)

    ws.cell(19, 6, "N° Accord SNTL")
    ws.cell(19, 6).font = _font(name=_SNTL_FONT, size=11)
    ws.cell(19, 6).alignment = _align(wrap=True)
    _style_range(ws, 19, 6, 19, 7, border=_border())

    article_lines, labor_total = _split_sntl_lines(lignes if lignes is not None else devis.lignes)
    article_start = 32
    min_article_rows = 10
    article_rows_count = max(min_article_rows, len(article_lines))
    table_header_row = article_start - 1
    headers = ["Référence article", "Désignation Article", "Quantité", "PU HT", "Total HT"]
    columns = [2, 3, 5, 6, 7]
    for col, header in zip(columns, headers):
        cell = ws.cell(table_header_row, col, header)
        cell.font = _font(name=_SNTL_FONT, size=12)
        cell.alignment = _align("center")
        cell.border = _border()
    ws.merge_cells(start_row=table_header_row, start_column=3, end_row=table_header_row, end_column=4)
    _style_range(ws, table_header_row, 2, table_header_row, 7, border=_border())

    article_total = Decimal("0.00")
    for offset in range(article_rows_count):
        row = article_start + offset
        ws.row_dimensions[row].height = 15.75
        for col in range(2, 8):
            ws.cell(row, col).border = _border()
            ws.cell(row, col).alignment = _align("center", wrap=True)
            ws.cell(row, col).font = _font(name=_SNTL_FONT, size=12)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        if offset >= len(article_lines):
            continue
        ligne = article_lines[offset]
        article_total += Decimal(str(ligne.total_ht or 0))
        ws.cell(row, 2, f"REF - N{offset + 1:03d}")
        ws.cell(row, 3, ligne.designation.upper())
        ws.cell(row, 5, _money(ligne.quantite))
        ws.cell(row, 6, _money(ligne.prix_unitaire_ht))
        ws.cell(row, 7, _money(ligne.total_ht))
        ws.cell(row, 5).number_format = "0.##"
        ws.cell(row, 6).number_format = '#,##0.00'
        ws.cell(row, 7).number_format = '#,##0.00'

    totals_start = article_start + article_rows_count
    total_ht = (article_total + labor_total).quantize(Decimal("0.01"))
    montant_tva = (total_ht * tva).quantize(Decimal("0.01"))
    montant_ttc = (total_ht + montant_tva).quantize(Decimal("0.01"))
    commission_sntl = calculer_commission_sntl(total_ht, montant_ttc)

    _sntl_total_row(ws, totals_start, "Montant Total article HT", article_total)
    _sntl_total_row(ws, totals_start + 1, "Main d'œuvre*", labor_total, label_end_col=4)
    _sntl_total_row(ws, totals_start + 2, "Montant Total article HT (1)", total_ht)
    _sntl_total_row(ws, totals_start + 3, f"TVA {int(tva * 100)}% (2)", montant_tva, rate=tva)
    _sntl_total_row(ws, totals_start + 4, "Montant Total TTC (1+2) (3)", montant_ttc)
    if include_commission:
        _sntl_total_row(ws, totals_start + 5, f"Commission SNTL (1x{int(commission * 100)}%) (4)", commission_sntl.commission_ht, rate=commission)
        _sntl_total_row(ws, totals_start + 6, f"TVA {int(tva * 100)}% sur la commission (4x{int(tva * 100)}%) (5)", commission_sntl.tva_commission, rate=tva)
        _sntl_total_row(ws, totals_start + 7, "Montant Net à régler (3-4-5)", commission_sntl.net_a_regler, bold_value=True)

    phrase_row = totals_start + (9 if include_commission else 7)
    phrase_subject = "la présente facture" if "FACTURE" in titre.upper() else "le présent devis"
    ws.cell(phrase_row, 2, f"Arrêté {phrase_subject} à la somme de {_capitalize(_amount_to_french(montant_ttc))} TTC")
    ws.cell(phrase_row, 2).alignment = _align(wrap=True)
    ws.cell(phrase_row, 2).font = _font(name=_SNTL_FONT, size=12)

    signature_row = phrase_row + 3
    ws.cell(signature_row, 5, "Cachet et signature")
    ws.cell(signature_row, 5).font = _font(name=_SNTL_FONT, size=12)
    ws.cell(signature_row, 5).alignment = _align()

    footnote_row = phrase_row + 8
    ws.cell(footnote_row, 2, "* Non valable pour les bons ateliers ''A\"")
    ws.cell(footnote_row, 2).font = _font(name=_SNTL_FONT, size=12)

    ws.print_area = f"A1:G{footnote_row}"


def _build_document_reparation(
    ws,
    *,
    titre: str,
    numero: str,
    date_document,
    devis: DevisReparation,
    statut: str,
    include_commission_sntl: bool,
    hide_etat: bool,
    montant_regle=None,
    lignes=None,
    montant_ht=None,
    montant_tva=None,
    montant_ttc=None,
) -> None:
    dossier = devis.dossier
    client = dossier.client
    vehicule = dossier.vehicule
    entreprise = obtenir_entreprise()

    _build_header(ws, titre, numero, date_document)

    ws.merge_cells("A6:C6")
    ws["A6"].value = "ATELIER"
    ws.merge_cells("E6:G6")
    ws["E6"].value = "CLIENT"
    for cell in (ws["A6"], ws["E6"]):
        cell.fill = _fill(_SLATE_950)
        cell.font = _font(bold=True, color=_JAUNE)
        cell.alignment = _align("center")
        cell.border = _border()

    atelier = [
        ("Raison sociale", entreprise.raison_sociale),
        ("Adresse", entreprise.adresse),
        ("Ville", entreprise.ville),
        ("Téléphone", entreprise.telephones),
        ("ICE", entreprise.ice),
        ("RC / IF", _join_parts(entreprise.rc, entreprise.if_fiscal, sep=" / ")),
        ("RIB", entreprise.rib),
    ]
    client_infos = [
        ("Nom", client.nom),
        ("Type", client.type_libelle),
        ("Téléphone", client.telephone),
        ("Adresse", client.adresse),
        ("Ville", client.ville),
        ("ICE", client.ice),
        ("Administration", client.administration_rattachee),
    ]
    for row_offset, (label, value) in enumerate(atelier, 7):
        _label_value(ws, row_offset, 1, label, value, width=3)
    for row_offset, (label, value) in enumerate(client_infos, 7):
        _label_value(ws, row_offset, 5, label, value, width=3)

    vehicle_row = 16
    ws.merge_cells(start_row=vehicle_row, start_column=1, end_row=vehicle_row, end_column=7)
    ws.cell(vehicle_row, 1, "VÉHICULE").fill = _fill(_SLATE_950)
    ws.cell(vehicle_row, 1).font = _font(bold=True, color=_JAUNE)
    ws.cell(vehicle_row, 1).alignment = _align("center")
    _style_range(ws, vehicle_row, 1, vehicle_row, 7, border=_border())

    vehicle_infos = [
        ("Matricule", vehicule.immatriculation),
        ("Marque et modèle", f"{vehicule.marque} {vehicule.modele}"),
        ("Kilométrage", dossier.kilometrage_entree or vehicule.kilometrage_actuel),
        ("Dossier", dossier.numero),
        ("Statut", statut),
        ("Assurance", dossier.assurance_nom),
    ]
    for col_index, (label, value) in enumerate(vehicle_infos, 1):
        row = 17 + ((col_index - 1) // 3)
        col = 1 + ((col_index - 1) % 3) * 2
        _label_value(ws, row, col, label, value, width=2)

    table_row = 21
    if hide_etat:
        headers = ["REFERENCE", "DESIGNATION", "QTE", "PU HT", "TOTAL HT"]
        table_columns = [1, 2, 5, 6, 7]
    else:
        headers = ["REFERENCE", "DESIGNATION", "ETAT", "QTE", "PU HT", "TOTAL HT"]
        table_columns = [1, 2, 4, 5, 6, 7]
    for col, header in zip(table_columns, headers):
        cell = ws.cell(table_row, col, header)
        cell.fill = _fill(_SLATE_950)
        cell.font = _font(bold=True, color=_BLANC)
        cell.alignment = _align("center", wrap=True)
        cell.border = _border()
    designation_end_col = 4 if hide_etat else 3
    ws.merge_cells(start_row=table_row, start_column=2, end_row=table_row, end_column=designation_end_col)

    first_line_row = table_row + 1
    lignes = list(lignes if lignes is not None else devis.lignes)
    montant_ht = devis.montant_ht if montant_ht is None else montant_ht
    montant_tva = devis.montant_tva if montant_tva is None else montant_tva
    montant_ttc = devis.montant_ttc if montant_ttc is None else montant_ttc

    for index, ligne in enumerate(lignes, first_line_row):
        if hide_etat:
            values = [
                f"REF-{index - table_row:03d}",
                ligne.designation,
                _money(ligne.quantite),
                _money(ligne.prix_unitaire_ht),
                _money(ligne.total_ht),
            ]
        else:
            values = [
                f"REF-{index - table_row:03d}",
                ligne.designation,
                ligne.etat_piece_libelle,
                _money(ligne.quantite),
                _money(ligne.prix_unitaire_ht),
                _money(ligne.total_ht),
            ]
        for col, value in zip(table_columns, values):
            cell = ws.cell(index, col, value)
            cell.border = _border("FFCBD5E1")
            cell.alignment = _align("center" if col != 2 else "left", wrap=True)
            if (hide_etat and col == 5) or (not hide_etat and col == 5):
                cell.number_format = "0.##"
            if col in {6, 7}:
                cell.number_format = '#,##0.00 "MAD"'
        ws.merge_cells(start_row=index, start_column=2, end_row=index, end_column=designation_end_col)

    last_line_row = max(first_line_row, first_line_row + len(lignes) - 1)
    totals_row = last_line_row + 2
    _totals_block(
        ws,
        totals_row,
        montant_ht,
        montant_tva,
        montant_ttc,
        montant_regle=montant_regle,
        include_commission_sntl=include_commission_sntl,
    )

    note_row = totals_row + (10 if include_commission_sntl else 7)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
    ws.cell(note_row, 1, _arrete(float(montant_ttc))).alignment = _align("left", wrap=True)
    ws.cell(note_row, 1).font = _font(bold=True, size=10)

    if devis.notes:
        ws.merge_cells(start_row=note_row + 2, start_column=1, end_row=note_row + 3, end_column=7)
        ws.cell(note_row + 2, 1, f"Notes: {devis.notes}").alignment = _align("left", wrap=True)
        _style_range(ws, note_row + 2, 1, note_row + 3, 7, fill=_fill(_SLATE_100), border=_border("FFCBD5E1"))

    ws.merge_cells(start_row=note_row + 5, start_column=5, end_row=note_row + 5, end_column=7)
    ws.cell(note_row + 5, 5, "Cachet et signature").font = _font(bold=True)
    ws.cell(note_row + 5, 5).alignment = _align("center")

    _set_print_options(ws)


def _totals_block(
    ws,
    start_row: int,
    montant_ht,
    montant_tva,
    montant_ttc,
    *,
    include_commission_sntl: bool,
    montant_regle=None,
) -> None:
    tva = _param_percent("taux_tva", Decimal("20"))
    commission = _param_percent("taux_commission_sntl", Decimal("10"))
    montant_ht = Decimal(str(montant_ht or 0)).quantize(Decimal("0.01"))
    montant_tva = Decimal(str(montant_tva or 0)).quantize(Decimal("0.01"))
    montant_ttc = Decimal(str(montant_ttc or 0)).quantize(Decimal("0.01"))
    commission_sntl = calculer_commission_sntl(montant_ht, montant_ttc)

    rows = [
        ("Montant HT", _money(montant_ht), _SLATE_100),
        (f"TVA {int(tva * 100)}%", _money(montant_tva), _SLATE_100),
        ("Montant TTC", _money(montant_ttc), _AMBRE_PALE),
    ]
    if montant_regle is not None:
        montant_regle = Decimal(str(montant_regle or 0)).quantize(Decimal("0.01"))
        reste = max(montant_ttc - montant_regle, Decimal("0.00"))
        rows.extend(
            [
                ("Montant encaissé", _money(montant_regle), _VERT_PALE),
                ("Reste à payer", _money(reste), _ROUGE_PALE if reste else _VERT_PALE),
            ]
        )
    if include_commission_sntl:
        rows.extend(
            [
                (f"Commission SNTL {int(commission * 100)}%", _money(commission_sntl.commission_ht), _ROUGE_PALE),
                (f"TVA {int(tva * 100)}% sur commission", _money(commission_sntl.tva_commission), _ROUGE_PALE),
                ("Montant net à régler", _money(commission_sntl.net_a_regler), _VERT_PALE),
            ]
        )

    for index, (label, value, fill_color) in enumerate(rows, start_row):
        ws.merge_cells(start_row=index, start_column=1, end_row=index, end_column=6)
        ws.cell(index, 1, label)
        ws.cell(index, 7, value)
        _style_range(ws, index, 1, index, 7, fill=_fill(fill_color), border=_border(), alignment=_align("right"))
        ws.cell(index, 1).font = _font(bold=True)
        ws.cell(index, 7).font = _font(bold=True)
        ws.cell(index, 7).number_format = '#,##0.00 "MAD"'


def _build_header(ws, titre: str, numero: str, date_document) -> None:
    for row in range(1, 5):
        ws.row_dimensions[row].height = 24

    ws.merge_cells("A1:B4")
    _style_range(ws, 1, 1, 4, 2, fill=_fill(_SLATE_950), border=_border(), alignment=_align("center"))
    _add_logo(ws)

    ws.merge_cells("C1:G2")
    ws["C1"].value = "MONSTER GARAGE"
    ws["C1"].font = _font(bold=True, size=24, color=_JAUNE)
    ws["C1"].alignment = _align("center")
    _style_range(ws, 1, 3, 2, 7, fill=_fill(_SLATE_950), border=_border(), alignment=_align("center"))

    ws.merge_cells("C3:G3")
    ws["C3"].value = "WIDINE MOTORS SERVICES"
    ws["C3"].font = _font(bold=True, size=13, color=_BLANC)
    ws["C3"].alignment = _align("center")
    _style_range(ws, 3, 3, 3, 7, fill=_fill(_SLATE_950), border=_border(), alignment=_align("center"))

    ws.merge_cells("C4:E4")
    ws["C4"].value = titre
    ws["C4"].font = _font(bold=True, size=14, color=_SLATE_950)
    ws["C4"].alignment = _align("center")
    ws["F4"].value = "N"
    ws["G4"].value = numero
    for cell in ("C4", "F4", "G4"):
        ws[cell].fill = _fill(_AMBRE_PALE)
        ws[cell].border = _border()
        ws[cell].alignment = _align("center")
        ws[cell].font = _font(bold=True)

    ws["F5"].value = "Date"
    ws["G5"].value = _date_value(date_document)
    ws["G5"].number_format = "dd/mm/yyyy"
    for cell in ("F5", "G5"):
        ws[cell].border = _border()
        ws[cell].alignment = _align("center")
        ws[cell].font = _font(bold=True)


def _write_table_header(ws, row: int, headers: list[str], *, start_col: int = 1) -> None:
    for offset, value in enumerate(headers, start_col):
        cell = ws.cell(row, offset, value)
        cell.fill = _fill(_SLATE_950)
        cell.font = _font(bold=True, color=_BLANC)
        cell.alignment = _align("center", wrap=True)
        cell.border = _border()


def _document_setup(ws, widths=None) -> None:
    widths = widths or [16, 28, 18, 18, 12, 16, 18]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35


def _document_setup_sntl(ws) -> None:
    widths = [2.71, 20.14, 28.43, 19.29, 21.29, 21.0, 18.14, 0.29, 11.57]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.scale = 62
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


def _build_sntl_official_header(ws, titre: str, numero: str, date_document) -> None:
    ws.merge_cells("A1:H3")
    ws["A1"].value = "MONSTER GARAGE"
    ws["A1"].font = _font(name=_SNTL_FONT, bold=True, italic=True, size=48)
    ws["A1"].alignment = _align("left")
    ws.row_dimensions[3].height = 36.75
    medium = Side(style="medium", color=_NOIR)
    for cell in ws[3][0:8]:
        cell.border = Border(bottom=medium)

    label = "Facture N°" if "FACTURE" in titre.upper() else "Devis N°"
    date_label = "Date facture:" if "FACTURE" in titre.upper() else "Date devis:"
    ws["D11"].value = label
    ws["E11"].value = numero
    ws["F11"].value = date_label
    ws["G11"].value = _date_value(date_document)
    ws["G11"].number_format = "dd/mm/yyyy"
    ws.row_dimensions[11].height = 18.75
    for cell_ref in ("D11", "E11", "F11", "G11"):
        ws[cell_ref].font = _font(name=_SNTL_FONT, bold=True, size=14)
        ws[cell_ref].alignment = _align("center" if cell_ref in {"E11", "G11"} else "left")

    ws.merge_cells("D13:E13")
    ws["D13"].value = "A"
    ws["D13"].font = _font(name=_SNTL_FONT, size=11)
    ws["D13"].alignment = _align("center")

    ws.merge_cells("C15:F15")
    ws["C15"].value = "La Société Nationale des Transports et de la Logistique (SNTL)"
    ws["C15"].font = _font(name=_SNTL_FONT, bold=True, size=18)
    ws["C15"].alignment = _align("center")
    ws.row_dimensions[15].height = 24


def _build_sntl_releve_header(ws) -> None:
    entreprise = obtenir_entreprise()

    ws.merge_cells("A1:H3")
    ws["A1"].value = "MONSTER GARAGE"
    ws["A1"].font = _font(name=_SNTL_FONT, bold=True, italic=True, size=48)
    ws["A1"].alignment = _align("left")
    ws.row_dimensions[3].height = 36.75
    medium = Side(style="medium", color=_NOIR)
    for cell in ws[3][0:8]:
        cell.border = Border(bottom=medium)

    ws.merge_cells("D11:E11")
    ws["D11"].value = "Relevé des Factures"
    ws["F11"].value = "Date:"
    ws["G11"].value = date.today()
    ws["G11"].number_format = "dd/mm/yyyy"
    ws.row_dimensions[11].height = 18.75
    for cell_ref in ("D11", "F11", "G11"):
        ws[cell_ref].font = _font(name=_SNTL_FONT, bold=True, size=14)
        ws[cell_ref].alignment = _align("center" if cell_ref in {"D11", "G11"} else "left")

    ws.merge_cells("D13:E13")
    ws["D13"].value = "A"
    ws["D13"].font = _font(name=_SNTL_FONT, size=11)
    ws["D13"].alignment = _align("center")

    ws.merge_cells("C15:F15")
    ws["C15"].value = "La Société Nationale des Transports et de la Logistique (SNTL)"
    ws["C15"].font = _font(name=_SNTL_FONT, bold=True, size=18)
    ws["C15"].alignment = _align("center")
    ws.row_dimensions[15].height = 24

    _sntl_section_header(ws, "B18:C18", "Partenaire")
    partenaire_rows = [
        ("N° Agrément SNTL", entreprise.agrement_sntl),
        ("Raison Sociale", entreprise.raison_sociale),
        ("Adresse", entreprise.adresse),
        ("Ville", entreprise.ville),
        ("RC", entreprise.rc),
        ("Patente", entreprise.patente),
        ("IF", entreprise.if_fiscal),
        ("ICE", entreprise.ice),
        ("N° RIB", entreprise.rib),
    ]
    for row, (label, value) in enumerate(partenaire_rows, 19):
        _sntl_label_value(ws, row, 2, label, value)


def _sntl_section_header(ws, cell_range: str, title: str) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = title
    cell.font = _font(name=_SNTL_FONT, bold=True, size=12)
    cell.alignment = _align("center")
    cell.parent.row_dimensions[cell.row].height = 15.75
    for row in ws[cell_range]:
        for item in row:
            item.border = _border()


def _sntl_label_value(ws, row: int, col: int, label: str, value) -> None:
    ws.cell(row, col, label)
    ws.cell(row, col + 1, _display(value))
    ws.cell(row, col).font = _font(name=_SNTL_FONT, size=11)
    ws.cell(row, col + 1).font = _font(name=_SNTL_FONT, bold=True, size=11)
    ws.cell(row, col).alignment = _align("left", wrap=True)
    ws.cell(row, col + 1).alignment = _align("left", wrap=True)
    _style_range(ws, row, col, row, col + 1, border=_border())


def _sntl_total_row(
    ws,
    row: int,
    label: str,
    value,
    *,
    rate: Decimal | None = None,
    bold_value: bool = False,
    label_end_col: int = 6,
) -> None:
    ws.row_dimensions[row].height = 15.75
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=label_end_col)
    ws.cell(row, 2, label)
    ws.cell(row, 7, _money(value))
    ws.cell(row, 2).alignment = _align("right" if label_end_col == 6 else "center")
    ws.cell(row, 7).alignment = _align("right")
    ws.cell(row, 7).number_format = '#,##0.00'
    ws.cell(row, 2).font = _font(name=_SNTL_FONT, size=12)
    ws.cell(row, 7).font = _font(name=_SNTL_FONT, bold=True, size=12)
    if bold_value:
        ws.cell(row, 7).font = _font(name=_SNTL_FONT, bold=True, size=12, color="FF0000FF")
    if rate:
        ws.cell(row, 9, float(rate))
        ws.cell(row, 9).number_format = "0%"
        ws.cell(row, 9).font = _font(name=_SNTL_FONT, size=12)
        ws.cell(row, 9).alignment = _align("center")
    _style_range(ws, row, 2, row, 7, border=_border())


def _split_sntl_lines(lines) -> tuple[list, Decimal]:
    article_lines = []
    labor_total = Decimal("0.00")
    for line in lines:
        if getattr(line, "type_ligne", None) == "main_oeuvre" or _is_labor_line(line.designation):
            labor_total += Decimal(str(line.total_ht or 0))
        else:
            article_lines.append(line)
    return article_lines, labor_total.quantize(Decimal("0.01"))


def _is_labor_line(designation: str) -> bool:
    normalized = (designation or "").lower().replace("\u0153", "oe").replace("Å“", "oe")
    return "main" in normalized and ("oeuvre" in normalized or "d'oeuvre" in normalized or "d oeuvre" in normalized)


def _format_sntl_matricule(value: object) -> str:
    raw = str(value or "").strip().upper()
    compact = re.sub(r"\s+", "", raw)
    match = re.fullmatch(r"([A-Z])[- ]?(\d+)", compact)
    if match:
        letter, number = match.groups()
        return f"{number.zfill(7)} - {letter}"

    match = re.fullmatch(r"(\d+)[- ]?([A-Z])", compact)
    if match:
        number, letter = match.groups()
        return f"{number.zfill(7)} - {letter}"

    return raw


def _sntl_bon_number(dossier, client) -> str:
    if getattr(dossier, "numero_bon_sntl", None):
        return dossier.numero_bon_sntl

    candidates = [dossier.notes, client.notes, dossier.numero]
    for candidate in candidates:
        text = str(candidate or "")
        match = re.search(r"(?:bon|or|ordre|n[°o])\D*(\d[\d\s-]*)", text, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return dossier.numero


def _commission_releve_sntl(facture: FactureReparation) -> Decimal:
    if facture.dossier.client.type != "sntl":
        return Decimal("0.00")
    return calculer_commission_sntl(facture.montant_ht, facture.montant_ttc).deduction_totale


def _label_value(ws, row: int, col: int, label: str, value, *, width: int) -> None:
    ws.cell(row, col, label)
    ws.cell(row, col + 1, _display(value))
    if width > 2:
        ws.merge_cells(start_row=row, start_column=col + 1, end_row=row, end_column=col + width - 1)
    ws.cell(row, col).font = _font(bold=True, color=_SLATE_500)
    ws.cell(row, col).alignment = _align("left", wrap=True)
    ws.cell(row, col + 1).alignment = _align("left", wrap=True)
    _style_range(ws, row, col, row, col + width - 1, border=_border("FFCBD5E1"))


def _add_logo(ws) -> None:
    if not _LOGO_PATH.exists():
        return
    logo = XLImage(str(_LOGO_PATH))
    logo.width = 150
    logo.height = 70
    logo.anchor = "A1"
    ws.add_image(logo)


def _set_print_options(ws) -> None:
    ws.print_title_rows = "1:5"


def _style_range(ws, min_row: int, min_col: int, max_row: int, max_col: int, *, fill=None, border=None, alignment=None) -> None:
    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in row:
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def _save(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _param_percent(cle: str, default: Decimal) -> Decimal:
    return param_percent(cle, default)


def _money(value) -> float:
    if value is None:
        return 0.0
    return float(Decimal(str(value)).quantize(Decimal("0.01")))


def _display(value) -> str:
    if value is None or value == "":
        return "-"
    return str(value)


def _date_value(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.today()


def _join_parts(*parts, sep=", ") -> str:
    return sep.join(str(part) for part in parts if part)


def _arrete(amount: float) -> str:
    return f"Arrêté le présent document à la somme de {amount:,.2f} dirhams TTC.".replace(",", " ")


def _amount_to_french(amount: Decimal) -> str:
    amount = Decimal(str(amount or 0)).quantize(Decimal("0.01"))
    dirhams = int(amount)
    centimes = int((amount - Decimal(dirhams)) * 100)
    words = f"{_number_to_french(dirhams)} dirhams"
    if centimes:
        words += f" et {_number_to_french(centimes)} centimes"
    return words


def _number_to_french(number: int) -> str:
    if number == 0:
        return "zero"

    units = [
        "",
        "un",
        "deux",
        "trois",
        "quatre",
        "cinq",
        "six",
        "sept",
        "huit",
        "neuf",
        "dix",
        "onze",
        "douze",
        "treize",
        "quatorze",
        "quinze",
        "seize",
    ]
    tens = {
        20: "vingt",
        30: "trente",
        40: "quarante",
        50: "cinquante",
        60: "soixante",
    }

    def below_hundred(value: int) -> str:
        if value < 17:
            return units[value]
        if value < 20:
            return f"dix {units[value - 10]}"
        if value < 70:
            ten = (value // 10) * 10
            rest = value % 10
            if rest == 0:
                return tens[ten]
            sep = " et " if rest == 1 else " "
            return f"{tens[ten]}{sep}{units[rest]}"
        if value < 80:
            rest = value - 60
            sep = " et " if rest == 11 else " "
            return f"soixante{sep}{below_hundred(rest)}"
        rest = value - 80
        if rest == 0:
            return "quatre vingt"
        return f"quatre vingt {below_hundred(rest)}"

    def below_thousand(value: int) -> str:
        if value < 100:
            return below_hundred(value)
        hundred = value // 100
        rest = value % 100
        prefix = "cent" if hundred == 1 else f"{units[hundred]} cent"
        return prefix if rest == 0 else f"{prefix} {below_hundred(rest)}"

    chunks = []
    millions = number // 1_000_000
    if millions:
        chunks.append(("un million" if millions == 1 else f"{below_thousand(millions)} millions"))
        number %= 1_000_000
    thousands = number // 1000
    if thousands:
        chunks.append("mille" if thousands == 1 else f"{below_thousand(thousands)} mille")
        number %= 1000
    if number:
        chunks.append(below_thousand(number))
    return " ".join(chunks)


def _capitalize(value: str) -> str:
    return value[:1].upper() + value[1:] if value else value


def _safe_sheet_title(value: str) -> str:
    return re.sub(r"[\[\]\:\*\?\/\\]", "-", value)[:31] or "DOCUMENT"


def _unique_sheet_title(wb: Workbook, value: str) -> str:
    base = _safe_sheet_title(value)[:31]
    if base not in wb.sheetnames:
        return base

    suffix = 2
    while True:
        candidate = f"{base[:28]} {suffix}"[:31]
        if candidate not in wb.sheetnames:
            return candidate
        suffix += 1


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("_")
    return cleaned or "document.xlsx"


def _font(*, bold=False, italic=False, size=10, color=_NOIR, name="Calibri") -> Font:
    return Font(name=name, bold=bold, italic=italic, size=size, color=color)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _border(color=_NOIR) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _align(horizontal="left", *, wrap=False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)
