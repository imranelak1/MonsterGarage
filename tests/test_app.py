from datetime import date, datetime, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook, load_workbook
import pytest

from app import create_app
from app.extensions import db
from app.models import AvanceSalaire, Client, DevisReparation, DossierReparation, Employe, Entreprise, FactureReparation, ParametreSysteme, Salaire, Utilisateur, Vehicule
from app.services.export_salaires import exporter_salaires_excel
from app.services.import_excel import importer_salaires_excel
from app.services.parametres import assurer_parametres_defaut, obtenir_entreprise


@pytest.fixture()
def app():
    app = create_app("testing")

    with app.app_context():
        db.create_all()
        obtenir_entreprise()
        assurer_parametres_defaut()
        utilisateur = Utilisateur(login="admin", nom_complet="Gérant Monster Garage", role="admin")
        utilisateur.definir_mot_de_passe("admin123")
        db.session.add(utilisateur)
        db.session.commit()

    yield app

    with app.app_context():
        db.session.remove()
        db.drop_all()


@pytest.fixture()
def client(app):
    return app.test_client()


def connecter(client):
    return client.post("/auth/connexion", data={"login": "admin", "mot_de_passe": "admin123"})


def assert_classeur_sans_sntl(wb):
    values = [
        cell.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]
    assert not any("SNTL" in value.upper() for value in values)


def creer_client_vehicule(app):
    with app.app_context():
        client_db = Client(code="FLOW1", type="particulier", nom="Client Workflow", telephone="0610000000")
        db.session.add(client_db)
        db.session.flush()
        vehicule = Vehicule(
            client_id=client_db.id,
            immatriculation="1111-F-6",
            marque="Toyota",
            modele="Hilux",
        )
        db.session.add(vehicule)
        db.session.commit()
        return client_db.id, vehicule.id


def creer_dossier_termine(client, app):
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "existant",
            "client_id": str(client_id),
            "vehicule_id": str(vehicule_id),
            "demande_client": "Finition et facturation",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={"objet": "Facturation finale", "designation_1": "Contrôle final", "quantite_1": "1", "prix_unitaire_ht_1": "200"},
    )
    with app.app_context():
        devis_id = DossierReparation.query.first().dernier_devis.id

    client.post(f"/dossiers/devis/{devis_id}/approuver", data={"mode_accord": "telephone"})
    client.post(f"/dossiers/{dossier_id}/terminer")
    return dossier_id


def test_application_demarre(client):
    response = client.get("/")

    assert response.status_code == 302
    assert "/tableau-de-bord" in response.location


def test_tableau_de_bord_demande_connexion(client):
    response = client.get("/tableau-de-bord")

    assert response.status_code == 302
    assert "/auth/connexion" in response.location


def test_tableau_de_bord_connecte_affiche_flux(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "existant",
            "client_id": str(client_id),
            "vehicule_id": str(vehicule_id),
            "demande_client": "Bruit moteur",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )

    response = client.get("/tableau-de-bord")

    assert response.status_code == 200
    assert "Démarrer un".encode() in response.data
    assert "Flux atelier en temps réel".encode() in response.data
    assert b"DA-" in response.data


def test_layout_connecte_affiche_user_topbar_et_toggle_sidebar(client):
    connecter(client)

    response = client.get("/tableau-de-bord")

    assert response.status_code == 200
    assert b"data-user-topbar" in response.data
    assert b"data-sidebar-toggle" in response.data
    assert "Gérant Monster Garage".encode() in response.data
    assert "Plein écran".encode() in response.data


def test_connexion_refuse_mauvais_mot_de_passe(client):
    response = client.post(
        "/auth/connexion",
        data={"login": "admin", "mot_de_passe": "mauvais"},
    )

    assert response.status_code == 200
    assert "Identifiant ou mot de passe incorrect".encode() in response.data


def test_page_connexion_design_atelier(client):
    response = client.get("/auth/connexion")

    assert response.status_code == 200
    assert b"id=\"login-shell\"" in response.data
    assert "Système de Gestion Atelier".encode() in response.data
    assert "Accès atelier".encode() in response.data
    assert b"data-login-theme" in response.data
    assert "Entrer dans l'atelier".encode() in response.data


def test_connexion_accepte_identifiants_valides(client):
    response = client.post(
        "/auth/connexion",
        data={"login": "admin", "mot_de_passe": "admin123"},
    )

    assert response.status_code == 302
    assert "/tableau-de-bord" in response.location


def test_deconnexion_termine_session(client):
    connecter(client)

    response = client.post("/auth/deconnexion")

    assert response.status_code == 302
    assert "/auth/connexion" in response.location


def test_session_expire_apres_inactivite(client):
    connecter(client)
    ancienne_activite = datetime.now(timezone.utc) - timedelta(minutes=31)

    with client.session_transaction() as session:
        session["derniere_activite"] = ancienne_activite.isoformat()

    response = client.get("/tableau-de-bord")

    assert response.status_code == 302
    assert "/auth/connexion" in response.location


def test_page_entreprise_reservee_au_gerant(client):
    response = client.get("/parametres/entreprise")

    assert response.status_code == 302
    assert "/auth/connexion" in response.location


def test_gerant_peut_modifier_entreprise(client, app):
    connecter(client)

    response = client.post(
        "/parametres/entreprise",
        data={
            "raison_sociale": "WIDINE MOTORS SERVICES",
            "nom_commercial": "MONSTER GARAGE",
            "adresse": "Quartier Industriel Sidi Ghanem",
            "ville": "Marrakech",
            "telephones": "0661 10 90 31",
            "email": "contact@monster-garage.test",
            "rc": "150241",
            "if_fiscal": "65976431",
            "ice": "003524622000063",
            "patente": "64006750",
            "cnss": "",
            "rib": "007450000612500000095129",
            "agrement_sntl": "3108",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        entreprise = Entreprise.query.first()
        assert entreprise.email == "contact@monster-garage.test"


def test_gerant_peut_modifier_parametres_systeme(client, app):
    connecter(client)

    response = client.post(
        "/parametres/systeme",
        data={
            "taux_tva": "20",
            "taux_commission_sntl": "10",
            "delai_paiement_default": "45",
            "charges_fixes_mensuelles": "70000",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        parametre = ParametreSysteme.query.filter_by(cle="delai_paiement_default").first()
        assert parametre.valeur == "45"


def test_liste_clients_demande_connexion(client):
    response = client.get("/clients/")

    assert response.status_code == 302
    assert "/auth/connexion" in response.location


def test_creation_client_avec_vehicule(client, app):
    connecter(client)

    response = client.post(
        "/clients/nouveau",
        data={
            "code": "C001",
            "type": "particulier",
            "nom": "Ahmed El Fassi",
            "sigle": "",
            "telephone": "0611111111",
            "telephone_2": "",
            "email": "ahmed@example.test",
            "adresse": "Marrakech",
            "ville": "Marrakech",
            "ice": "",
            "if_fiscal": "",
            "rc": "",
            "administration_rattachee": "",
            "delai_paiement_jours": "30",
            "notes": "",
            "immatriculation": "12345-A-6",
            "marque": "Dacia",
            "modele": "Logan",
            "type_immatriculation": "standard",
            "type_vehicule": "voiture",
            "type_carburant": "diesel",
            "kilometrage_actuel": "120000",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        client_cree = Client.query.filter_by(code="C001").first()
        assert client_cree is not None
        assert client_cree.vehicules[0].immatriculation == "12345-A-6"


def test_creation_client_refuse_telephone_invalide(client, app):
    connecter(client)

    response = client.post(
        "/clients/nouveau",
        data={
            "code": "C009",
            "type": "particulier",
            "nom": "Client Téléphone Faux",
            "telephone": "123",
            "telephone_2": "",
            "email": "",
            "adresse": "",
            "ville": "Marrakech",
            "ice": "",
            "if_fiscal": "",
            "rc": "",
            "administration_rattachee": "",
            "delai_paiement_jours": "30",
            "notes": "",
            "immatriculation": "",
            "marque": "",
            "modele": "",
        },
    )

    assert response.status_code == 200
    assert "Numéro de téléphone invalide".encode() in response.data
    with app.app_context():
        assert Client.query.filter_by(code="C009").first() is None


def test_recherche_client(client, app):
    connecter(client)
    with app.app_context():
        db.session.add(Client(code="ADM1", type="administration", nom="Wilaya de Marrakech"))
        db.session.commit()

    response = client.get("/clients/?q=wilaya")

    assert response.status_code == 200
    assert "Wilaya de Marrakech".encode() in response.data


def test_ajout_vehicule_depuis_fiche_client(client, app):
    connecter(client)
    with app.app_context():
        client_db = Client(code="C002", type="particulier", nom="Atlas Transport")
        db.session.add(client_db)
        db.session.commit()
        client_id = client_db.id

    response = client.post(
        f"/clients/{client_id}/vehicules/nouveau",
        data={
            "immatriculation": "9988-B-6",
            "marque": "Renault",
            "modele": "Master",
            "type_vehicule": "utilitaire",
            "type_carburant": "diesel",
            "kilometrage_actuel": "90000",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        vehicule = Vehicule.query.filter_by(immatriculation="9988-B-6").first()
        assert vehicule is not None
        assert vehicule.client.nom == "Atlas Transport"


def test_liste_vehicules(client, app):
    connecter(client)
    with app.app_context():
        client_db = Client(code="C003", type="particulier", nom="Salma Amrani")
        db.session.add(client_db)
        db.session.flush()
        db.session.add(
            Vehicule(
                client_id=client_db.id,
                immatriculation="7777-D-6",
                marque="Peugeot",
                modele="208",
            )
        )
        db.session.commit()

    response = client.get("/vehicules/?q=7777")

    assert response.status_code == 200
    assert b"7777-D-6" in response.data
    assert b"Peugeot" in response.data


def test_creation_dossier_atelier(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)

    response = client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "existant",
            "client_id": str(client_id),
            "vehicule_id": str(vehicule_id),
            "demande_client": "Bruit moteur",
            "diagnostic_initial": "Contrôle à faire",
            "kilometrage_entree": "150000",
            "notes": "",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        dossier = DossierReparation.query.first()
        assert dossier.statut == "pending_devis"
        assert dossier.numero.startswith("DA-")


def test_journal_actions_affiche_date(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)

    client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "existant",
            "client_id": str(client_id),
            "vehicule_id": str(vehicule_id),
            "demande_client": "Bruit moteur",
            "diagnostic_initial": "Contrôle à faire",
            "kilometrage_entree": "150000",
            "notes": "",
        },
    )

    with app.app_context():
        dossier = DossierReparation.query.first()
        dossier_id = dossier.id
        date_action = dossier.journal[0].created_at.strftime("%d/%m/%Y %H:%M")

    response = client.get(f"/dossiers/{dossier_id}")

    assert response.status_code == 200
    assert date_action.encode() in response.data
    assert b"datetime=" in response.data


def test_detail_dossier_priorise_poste_de_commande(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "existant",
            "client_id": str(client_id),
            "vehicule_id": str(vehicule_id),
            "demande_client": "Contrôle workflow prioritaire",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    response = client.get(f"/dossiers/{dossier_id}")

    assert response.status_code == 200
    poste = response.data.index("Poste de commande".encode())
    devis = response.data.index("Devis versionnés".encode())
    journal = response.data.index("Journal d'actions".encode())
    assert poste < devis < journal
    assert "Prochaine action".encode() in response.data


def test_formulaire_dossier_prepare_filtrage_vehicules_par_client(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)

    response = client.get("/dossiers/nouveau")

    assert response.status_code == 200
    assert f'data-client-id="{client_id}"'.encode() in response.data
    assert f'value="{vehicule_id}"'.encode() in response.data
    assert b"data-client-type=\"particulier\"" in response.data
    assert b"data-assurance-block" in response.data
    assert b"600 - Commune Harbil / OR 85" in response.data
    assert b"100 - Wilaya / OR 83" in response.data
    assert b"Nouveau client SNTL" in response.data
    assert "Sélectionnez un client existant pour afficher uniquement ses véhicules.".encode() in response.data


def test_creation_dossier_particulier_enregistre_assurance(client, app):
    connecter(client)

    response = client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "nouveau",
            "client_nom": "Client Assuré",
            "client_telephone": "0622222222",
            "client_type": "particulier",
            "assurance_nom": "Wafa Assurance",
            "client_code": "",
            "mode_vehicule": "nouveau",
            "vehicule_immatriculation": "1010-A-6",
            "vehicule_marque": "Renault",
            "vehicule_modele": "Clio",
            "vehicule_type": "voiture",
            "vehicule_carburant": "diesel",
            "vehicule_type_immatriculation": "standard",
            "demande_client": "Déclaration assurance",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        dossier = DossierReparation.query.first()
        assert dossier.assurance_nom == "Wafa Assurance"


def test_creation_dossier_sntl_force_immatriculation_administrative(client, app):
    connecter(client)
    with app.app_context():
        client_db = Client(code="SNTL2", type="sntl", nom="Client SNTL Admin", telephone="0610000000")
        db.session.add(client_db)
        db.session.commit()
        client_id = client_db.id

    response = client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "client_id": str(client_id),
            "mode_vehicule": "nouveau",
            "vehicule_immatriculation": "2020-S-6",
            "vehicule_marque": "Ford",
            "vehicule_modele": "Ranger",
            "vehicule_type": "utilitaire",
            "vehicule_carburant": "diesel",
            "vehicule_type_immatriculation": "standard",
            "demande_client": "Dossier SNTL",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        dossier = DossierReparation.query.first()
        vehicule = Vehicule.query.filter_by(immatriculation="2020-S-6").first()
        assert dossier.numero_bon_sntl.isdigit()
        assert len(dossier.numero_bon_sntl) == 12
        assert vehicule.type_immatriculation == "administrative"


def test_creation_dossier_sntl_predefini_cree_client_code_or(client, app):
    connecter(client)

    response = client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "nouveau",
            "client_type": "sntl",
            "client_sntl_preset": "600-85",
            "client_nom": "",
            "client_code": "",
            "client_ice": "ICE600HAR",
            "client_ville": "Marrakech",
            "mode_vehicule": "nouveau",
            "vehicule_immatriculation": "J207789",
            "vehicule_marque": "Ford",
            "vehicule_modele": "Transit",
            "vehicule_type": "utilitaire",
            "vehicule_carburant": "diesel",
            "vehicule_type_immatriculation": "standard",
            "demande_client": "Ticket SNTL Harbil",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        client_db = Client.query.filter_by(code="600").first()
        dossier = DossierReparation.query.first()
        vehicule = Vehicule.query.filter_by(immatriculation="J207789").first()
        assert client_db is not None
        assert client_db.type == "sntl"
        assert client_db.nom == "Commune Harbil"
        assert client_db.ice == "ICE600HAR"
        assert client_db.notes == "OR numero 85"
        assert dossier.client_id == client_db.id
        assert dossier.numero_bon_sntl.isdigit()
        assert len(dossier.numero_bon_sntl) == 12
        assert vehicule.type_immatriculation == "administrative"


def test_creation_dossier_sntl_predefini_reutilise_client_existant(client, app):
    connecter(client)
    with app.app_context():
        client_db = Client(code="100", type="sntl", nom="Wilaya", notes="OR numero 83")
        db.session.add(client_db)
        db.session.commit()
        client_id = client_db.id

    response = client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "nouveau",
            "client_type": "sntl",
            "client_sntl_preset": "100-83",
            "mode_vehicule": "nouveau",
            "vehicule_immatriculation": "J100083",
            "vehicule_marque": "Dacia",
            "vehicule_modele": "Duster",
            "vehicule_type": "voiture",
            "vehicule_carburant": "diesel",
            "vehicule_type_immatriculation": "standard",
            "demande_client": "Ticket Wilaya",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        assert Client.query.filter_by(code="100").count() == 1
        assert DossierReparation.query.first().client_id == client_id


def test_creation_dossier_sntl_custom_reste_possible(client, app):
    connecter(client)

    response = client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "nouveau",
            "client_type": "sntl",
            "client_sntl_preset": "custom",
            "client_nom": "Nouvelle Administration SNTL",
            "client_code": "900",
            "client_ville": "Marrakech",
            "mode_vehicule": "nouveau",
            "vehicule_immatriculation": "J900001",
            "vehicule_marque": "Renault",
            "vehicule_modele": "Master",
            "vehicule_type": "utilitaire",
            "vehicule_carburant": "diesel",
            "vehicule_type_immatriculation": "standard",
            "demande_client": "Nouveau client SNTL",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        client_db = Client.query.filter_by(code="900").first()
        assert client_db is not None
        assert client_db.type == "sntl"
        assert client_db.nom == "Nouvelle Administration SNTL"


def test_creation_dossier_cree_client_et_vehicule(client, app):
    connecter(client)

    response = client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "nouveau",
            "client_nom": "Nouveau Client Atelier",
            "client_telephone": "0622222222",
            "client_type": "particulier",
            "client_code": "",
            "client_email": "",
            "client_ville": "Marrakech",
            "client_adresse": "Sidi Ghanem",
            "mode_vehicule": "nouveau",
            "vehicule_immatriculation": "5555-H-6",
            "vehicule_marque": "Hyundai",
            "vehicule_modele": "Tucson",
            "vehicule_type": "voiture",
            "vehicule_carburant": "diesel",
            "vehicule_type_immatriculation": "standard",
            "demande_client": "Véhicule arrivé pour diagnostic",
            "diagnostic_initial": "À inspecter",
            "kilometrage_entree": "80000",
            "notes": "",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        client_db = Client.query.filter_by(nom="Nouveau Client Atelier").first()
        vehicule = Vehicule.query.filter_by(immatriculation="5555-H-6").first()
        dossier = DossierReparation.query.first()
        assert client_db is not None
        assert vehicule is not None
        assert dossier.client_id == client_db.id
        assert dossier.vehicule_id == vehicule.id


def test_recherche_dossiers_par_client_vehicule_et_matricule(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "existant",
            "client_id": str(client_id),
            "vehicule_id": str(vehicule_id),
            "demande_client": "Bruit moteur spécifique",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )

    response = client.get("/dossiers/?q=Hilux")
    assert response.status_code == 200
    assert b"1111-F-6" in response.data

    response = client.get("/dossiers/?q=Client+Workflow")
    assert response.status_code == 200
    assert "Client Workflow".encode() in response.data


def test_creation_dossier_client_existant_nouveau_vehicule_lie_au_client(client, app):
    connecter(client)
    client_id, _vehicule_id = creer_client_vehicule(app)

    response = client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "client_id": str(client_id),
            "mode_vehicule": "nouveau",
            "vehicule_immatriculation": "9090-MG-6",
            "vehicule_marque": "Volkswagen",
            "vehicule_modele": "Golf",
            "vehicule_type": "voiture",
            "vehicule_carburant": "diesel",
            "vehicule_type_immatriculation": "standard",
            "demande_client": "Nouvelle voiture du même client",
            "diagnostic_initial": "",
            "kilometrage_entree": "45000",
            "notes": "",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        vehicule = Vehicule.query.filter_by(immatriculation="9090-MG-6").first()
        dossier = DossierReparation.query.order_by(DossierReparation.id.desc()).first()
        assert vehicule is not None
        assert vehicule.client_id == client_id
        assert dossier.client_id == client_id
        assert dossier.vehicule_id == vehicule.id


def test_creation_dossier_refuse_telephone_invalide(client, app):
    connecter(client)

    response = client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "nouveau",
            "client_nom": "Client Téléphone Invalide",
            "client_telephone": "abc",
            "client_type": "particulier",
            "client_code": "",
            "mode_vehicule": "nouveau",
            "vehicule_immatriculation": "1212-H-6",
            "vehicule_marque": "Dacia",
            "vehicule_modele": "Duster",
            "demande_client": "Contrôle",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )

    assert response.status_code == 200
    assert "Numéro de téléphone invalide".encode() in response.data
    with app.app_context():
        assert DossierReparation.query.first() is None


def test_devis_initial_et_approbation(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "existant",
            "client_id": str(client_id),
            "vehicule_id": str(vehicule_id),
            "demande_client": "Freinage faible",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )

    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    response = client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={
            "objet": "Freins avant",
            "designation_1": "Plaquettes avant",
            "quantite_1": "1",
            "prix_unitaire_ht_1": "500",
            "designation_2": "Main d'oeuvre",
            "quantite_2": "2",
            "prix_unitaire_ht_2": "150",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        dossier = db.session.get(DossierReparation, dossier_id)
        devis = dossier.dernier_devis
        assert dossier.statut == "pending_approval"
        assert devis.version == 1
        assert str(devis.montant_ttc) == "960.00"
        devis_id = devis.id

    response = client.post(f"/dossiers/devis/{devis_id}/approuver", data={"mode_accord": "telephone"})

    assert response.status_code == 302
    with app.app_context():
        dossier = db.session.get(DossierReparation, dossier_id)
        devis = db.session.get(DevisReparation, devis_id)
        assert devis.statut == "approved"
        assert dossier.statut == "in_progress"


def test_devis_lignes_dynamiques_et_piece_occasion(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "existant",
            "client_id": str(client_id),
            "vehicule_id": str(vehicule_id),
            "demande_client": "Pare-chocs",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    response = client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={
            "objet": "Carrosserie",
            "designation": ["Pare-chocs occasion", "Peinture"],
            "quantite": ["1", "2"],
            "prix_unitaire_ht": ["700", "250"],
            "etat_piece": ["occasion", "neuf"],
        },
    )

    assert response.status_code == 302
    with app.app_context():
        devis = db.session.get(DossierReparation, dossier_id).dernier_devis
        assert len(devis.lignes) == 2
        assert devis.lignes[0].etat_piece == "occasion"
        assert str(devis.montant_ttc) == "1440.00"


def test_formulaire_devis_propose_ajout_main_oeuvre(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "existant",
            "client_id": str(client_id),
            "vehicule_id": str(vehicule_id),
            "demande_client": "Diagnostic main d'oeuvre",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    response = client.get(f"/dossiers/{dossier_id}/devis/nouveau")

    assert response.status_code == 200
    assert b"data-add-labor" in response.data
    assert "Ajouter main d'oeuvre".encode() in response.data
    assert "Coût main d'oeuvre HT".encode() in response.data


def test_devis_sntl_force_pieces_neuves(client, app):
    connecter(client)
    with app.app_context():
        client_db = Client(code="SNTL1", type="sntl", nom="Administration SNTL", telephone="0610000000")
        db.session.add(client_db)
        db.session.flush()
        vehicule = Vehicule(client_id=client_db.id, immatriculation="8888-S-6", marque="Ford", modele="Transit")
        db.session.add(vehicule)
        db.session.commit()
        client_id, vehicule_id = client_db.id, vehicule.id

    client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "existant",
            "client_id": str(client_id),
            "vehicule_id": str(vehicule_id),
            "demande_client": "Réparation flotte",
            "diagnostic_initial": "",
            "kilometrage_entree": "",
            "notes": "",
        },
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={
            "objet": "SNTL",
            "designation": ["Alternateur"],
            "quantite": ["1"],
            "prix_unitaire_ht": ["1800"],
            "etat_piece": ["occasion"],
        },
    )

    with app.app_context():
        ligne = db.session.get(DossierReparation, dossier_id).dernier_devis.lignes[0]
        assert ligne.etat_piece == "neuf"


def test_un_seul_devis_actif_a_la_fois(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={"mode_client": "existant", "mode_vehicule": "existant", "client_id": client_id, "vehicule_id": vehicule_id, "demande_client": "Diagnostic", "diagnostic_initial": "", "kilometrage_entree": "", "notes": ""},
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    donnees_devis = {
        "objet": "Diagnostic",
        "designation_1": "Recherche panne",
        "quantite_1": "1",
        "prix_unitaire_ht_1": "300",
    }
    client.post(f"/dossiers/{dossier_id}/devis/nouveau", data=donnees_devis)
    response = client.post(f"/dossiers/{dossier_id}/devis/nouveau", data=donnees_devis)

    assert response.status_code == 200
    assert "devis est".encode() in response.data


def test_boucle_devis_complementaire_et_blocage_ancienne_version(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={"mode_client": "existant", "mode_vehicule": "existant", "client_id": client_id, "vehicule_id": vehicule_id, "demande_client": "Révision", "diagnostic_initial": "", "kilometrage_entree": "", "notes": ""},
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={"objet": "Révision v1", "designation_1": "Vidange", "quantite_1": "1", "prix_unitaire_ht_1": "400"},
    )
    with app.app_context():
        devis_v1_id = DossierReparation.query.first().dernier_devis.id
    client.post(f"/dossiers/devis/{devis_v1_id}/approuver", data={"mode_accord": "presentiel"})

    client.post(f"/dossiers/{dossier_id}/pause", data={"raison": "Courroie à remplacer"})
    client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={"objet": "Révision v2", "designation_1": "Courroie", "quantite_1": "1", "prix_unitaire_ht_1": "900"},
    )

    response = client.post(f"/dossiers/devis/{devis_v1_id}/approuver", data={"mode_accord": "telephone"})

    assert response.status_code == 302
    with app.app_context():
        dossier = db.session.get(DossierReparation, dossier_id)
        assert dossier.dernier_devis.version == 2
        assert dossier.dernier_devis.statut == "pending"
        assert dossier.statut == "pending_approval"


def test_nouvelle_version_devis_reprend_anciennes_lignes(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={"mode_client": "existant", "mode_vehicule": "existant", "client_id": client_id, "vehicule_id": vehicule_id, "demande_client": "Révision", "diagnostic_initial": "", "kilometrage_entree": "", "notes": ""},
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={"objet": "Révision initiale", "designation": ["Vidange spéciale"], "quantite": ["1"], "prix_unitaire_ht": ["400"], "etat_piece": ["neuf"]},
    )
    with app.app_context():
        devis_id = DossierReparation.query.first().dernier_devis.id
    client.post(f"/dossiers/devis/{devis_id}/approuver", data={"mode_accord": "presentiel"})
    client.post(f"/dossiers/{dossier_id}/pause", data={"raison": "Filtre supplémentaire"})

    response = client.get(f"/dossiers/{dossier_id}/devis/nouveau")

    assert response.status_code == 200
    assert "Les lignes du devis v1 sont reprises".encode() in response.data
    assert "Vidange spéciale".encode() in response.data


def test_refus_devis_affiche_action_suivante(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={"mode_client": "existant", "mode_vehicule": "existant", "client_id": client_id, "vehicule_id": vehicule_id, "demande_client": "Diagnostic", "diagnostic_initial": "", "kilometrage_entree": "", "notes": ""},
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={"objet": "Diagnostic", "designation": ["Recherche panne"], "quantite": ["1"], "prix_unitaire_ht": ["300"], "etat_piece": ["neuf"]},
    )
    with app.app_context():
        devis_id = DossierReparation.query.first().dernier_devis.id

    response = client.post(f"/dossiers/devis/{devis_id}/refuser", data={"motif_refus": "Trop cher"}, follow_redirects=True)

    assert response.status_code == 200
    assert "Créez une version corrigée".encode() in response.data
    assert "Créer une version corrigée".encode() in response.data


def test_terminer_dossier_apres_dernier_devis_approuve(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={"mode_client": "existant", "mode_vehicule": "existant", "client_id": client_id, "vehicule_id": vehicule_id, "demande_client": "Finition", "diagnostic_initial": "", "kilometrage_entree": "", "notes": ""},
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={"objet": "Finition", "designation_1": "Contrôle", "quantite_1": "1", "prix_unitaire_ht_1": "200"},
    )
    with app.app_context():
        devis_id = DossierReparation.query.first().dernier_devis.id
    client.post(f"/dossiers/devis/{devis_id}/approuver", data={"mode_accord": "telephone"})

    response = client.post(f"/dossiers/{dossier_id}/terminer")

    assert response.status_code == 302
    with app.app_context():
        dossier = db.session.get(DossierReparation, dossier_id)
        assert dossier.statut == "completed"


def test_facture_finale_generee_depuis_dernier_devis_approuve(client, app):
    connecter(client)
    dossier_id = creer_dossier_termine(client, app)

    response = client.post(f"/factures/dossiers/{dossier_id}/generer")

    assert response.status_code == 302
    assert "/factures/" in response.location
    with app.app_context():
        facture = FactureReparation.query.first()
        dossier = db.session.get(DossierReparation, dossier_id)
        assert facture is not None
        assert facture.numero.startswith("FA-")
        assert facture.statut == "emise"
        assert str(facture.montant_ttc) == "240.00"
        assert facture.devis_id == dossier.dernier_devis_approuve.id


def test_facture_ne_peut_pas_etre_generee_avant_fin_reparation(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={"mode_client": "existant", "mode_vehicule": "existant", "client_id": client_id, "vehicule_id": vehicule_id, "demande_client": "Réparation", "diagnostic_initial": "", "kilometrage_entree": "", "notes": ""},
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    response = client.post(f"/factures/dossiers/{dossier_id}/generer", follow_redirects=True)

    assert response.status_code == 200
    assert b"apres la fin de la reparation" in response.data
    with app.app_context():
        assert FactureReparation.query.first() is None


def test_accord_devis_accepte_signature(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={"mode_client": "existant", "mode_vehicule": "existant", "client_id": client_id, "vehicule_id": vehicule_id, "demande_client": "Accord signe", "diagnostic_initial": "", "kilometrage_entree": "", "notes": ""},
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={"objet": "Accord signe", "designation_1": "Diagnostic", "quantite_1": "1", "prix_unitaire_ht_1": "150"},
    )
    with app.app_context():
        devis_id = DossierReparation.query.first().dernier_devis.id

    response = client.post(f"/dossiers/devis/{devis_id}/approuver", data={"mode_accord": "signature"})

    assert response.status_code == 302
    with app.app_context():
        devis = db.session.get(DevisReparation, devis_id)
        dossier = db.session.get(DossierReparation, dossier_id)
        assert devis.mode_accord == "signature"
        assert dossier.statut == "in_progress"


def test_annulation_facturable_genere_facture_travaux_effectues(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={"mode_client": "existant", "mode_vehicule": "existant", "client_id": client_id, "vehicule_id": vehicule_id, "demande_client": "Client annule apres diagnostic", "diagnostic_initial": "", "kilometrage_entree": "", "notes": ""},
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={"objet": "Travaux effectues", "designation_1": "Diagnostic realise", "quantite_1": "1", "prix_unitaire_ht_1": "200"},
    )
    with app.app_context():
        devis_id = DossierReparation.query.first().dernier_devis.id
    client.post(f"/dossiers/devis/{devis_id}/approuver", data={"mode_accord": "telephone"})

    response = client.post(f"/dossiers/{dossier_id}/annuler-facturable", data={"motif": "Client stoppe la reparation"})
    assert response.status_code == 302
    with app.app_context():
        dossier = db.session.get(DossierReparation, dossier_id)
        assert dossier.statut == "cancelled_billable"

    response = client.post(f"/factures/dossiers/{dossier_id}/generer")

    assert response.status_code == 302
    with app.app_context():
        facture = FactureReparation.query.first()
        assert facture is not None
        assert str(facture.montant_ttc) == "240.00"
        assert facture.dossier.statut == "cancelled_billable"


def test_annulation_facturable_exige_devis_approuve(client, app):
    connecter(client)
    client_id, vehicule_id = creer_client_vehicule(app)
    client.post(
        "/dossiers/nouveau",
        data={"mode_client": "existant", "mode_vehicule": "existant", "client_id": client_id, "vehicule_id": vehicule_id, "demande_client": "Annulation sans travaux chiffres", "diagnostic_initial": "", "kilometrage_entree": "", "notes": ""},
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id

    response = client.post(
        f"/dossiers/{dossier_id}/annuler-facturable",
        data={"motif": "Aucun devis approuve"},
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"devis limite aux travaux effectues" in response.data
    with app.app_context():
        dossier = db.session.get(DossierReparation, dossier_id)
        assert dossier.statut == "pending_devis"
        assert FactureReparation.query.first() is None


def test_reprise_garantie_continue_meme_dossier_apres_facture(client, app):
    connecter(client)
    dossier_id = creer_dossier_termine(client, app)
    client.post(f"/factures/dossiers/{dossier_id}/generer")
    with app.app_context():
        facture_id = FactureReparation.query.first().id
    client.post(f"/factures/{facture_id}/livrer")
    client.post(
        f"/factures/{facture_id}/regler",
        data={"mode_reglement": "virement", "montant_reglement": "240", "reference_reglement": "GAR-001"},
    )

    response = client.post(
        f"/dossiers/{dossier_id}/garantie/reouvrir",
        data={"motif": "Retour client sous garantie"},
    )

    assert response.status_code == 302
    with app.app_context():
        dossier = db.session.get(DossierReparation, dossier_id)
        assert dossier.statut == "in_progress"
        assert dossier.facture.id == facture_id
        assert FactureReparation.query.count() == 1
        assert any(action.action == "reprise_garantie" for action in dossier.journal)

    client.post(f"/dossiers/{dossier_id}/terminer")
    with app.app_context():
        dossier = db.session.get(DossierReparation, dossier_id)
        assert dossier.statut == "completed"
        assert FactureReparation.query.count() == 1


def test_facture_livraison_puis_reglement(client, app):
    connecter(client)
    dossier_id = creer_dossier_termine(client, app)
    client.post(f"/factures/dossiers/{dossier_id}/generer")
    with app.app_context():
        facture_id = FactureReparation.query.first().id

    response = client.post(
        f"/factures/{facture_id}/regler",
        data={"mode_reglement": "virement", "montant_reglement": "240", "reference_reglement": "VIR-001"},
        follow_redirects=True,
    )
    assert b"apres la livraison du vehicule" in response.data

    response = client.post(f"/factures/{facture_id}/livrer")
    assert response.status_code == 302
    response = client.post(
        f"/factures/{facture_id}/regler",
        data={"mode_reglement": "virement", "montant_reglement": "240", "reference_reglement": "VIR-001"},
    )

    assert response.status_code == 302
    with app.app_context():
        facture = db.session.get(FactureReparation, facture_id)
        assert facture.statut == "reglee"
        assert facture.mode_reglement == "virement"
        assert facture.reference_reglement == "VIR-001"
        assert str(facture.montant_regle) == "240.00"


def test_facture_accepte_paiements_partiels_sans_depassement(client, app):
    connecter(client)
    dossier_id = creer_dossier_termine(client, app)
    client.post(f"/factures/dossiers/{dossier_id}/generer")
    with app.app_context():
        facture_id = FactureReparation.query.first().id

    client.post(f"/factures/{facture_id}/livrer")
    response = client.post(
        f"/factures/{facture_id}/regler",
        data={"mode_reglement": "especes", "montant_reglement": "100", "reference_reglement": "ACOMPTE"},
    )
    assert response.status_code == 302
    with app.app_context():
        facture = db.session.get(FactureReparation, facture_id)
        assert facture.statut == "livree"
        assert facture.statut_libelle == "Partiellement reglee"
        assert str(facture.montant_regle) == "100.00"
        assert str(facture.montant_restant) == "140.00"
        client_id = facture.dossier.client_id

    export_response = client.get(f"/factures/{facture_id}/telecharger")
    wb = load_workbook(BytesIO(export_response.data), data_only=False)
    ws = wb.active
    assert "Montant encaisse" in [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert "Reste a payer" in [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert 100 in [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert 140 in [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    releve_response = client.get(f"/factures/clients/{client_id}/releve")
    releve = load_workbook(BytesIO(releve_response.data), data_only=False).active
    assert releve["E10"].value == 100
    assert releve["F10"].value == "=D10-E10"

    response = client.post(
        f"/factures/{facture_id}/regler",
        data={"mode_reglement": "especes", "montant_reglement": "200", "reference_reglement": "TROP"},
        follow_redirects=True,
    )
    assert "depasse le reste".encode() in response.data
    with app.app_context():
        facture = db.session.get(FactureReparation, facture_id)
        assert str(facture.montant_regle) == "100.00"

    client.post(
        f"/factures/{facture_id}/regler",
        data={"mode_reglement": "virement", "montant_reglement": "140", "reference_reglement": "SOLDE"},
    )
    with app.app_context():
        facture = db.session.get(FactureReparation, facture_id)
        assert facture.statut == "reglee"
        assert str(facture.montant_regle) == "240.00"
        assert str(facture.montant_restant) == "0.00"


def test_liste_factures_affiche_facture(client, app):
    connecter(client)
    dossier_id = creer_dossier_termine(client, app)
    client.post(f"/factures/dossiers/{dossier_id}/generer")

    response = client.get("/factures/")

    assert response.status_code == 200
    assert b"FA-" in response.data
    assert "Client Workflow".encode() in response.data


def test_export_devis_excel_depuis_dossier(client, app):
    connecter(client)
    dossier_id = creer_dossier_termine(client, app)
    with app.app_context():
        devis_id = db.session.get(DossierReparation, dossier_id).dernier_devis.id

    response = client.get(f"/dossiers/devis/{devis_id}/telecharger")

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    wb = load_workbook(BytesIO(response.data), data_only=False)
    assert_classeur_sans_sntl(wb)
    ws = wb.active
    assert ws["C1"].value == "MONSTER GARAGE"
    assert ws["C4"].value == "DEVIS DE REPARATION"
    assert ws["G4"].value.endswith("-V1")
    assert ws.freeze_panes is None
    assert ws["E22"].value == 1
    assert ws["E22"].number_format == "0.##"
    assert ws["F22"].value == 200
    assert ws["G22"].value == 200
    assert ws["G24"].value == 200
    assert ws["G25"].value == 40
    assert ws["G26"].value == 240
    assert any(isinstance(cell.value, str) and "final" in cell.value for row in ws.iter_rows() for cell in row)


def test_export_devis_pdf_vue_et_telechargement(client, app):
    connecter(client)
    dossier_id = creer_dossier_termine(client, app)
    with app.app_context():
        devis_id = db.session.get(DossierReparation, dossier_id).dernier_devis.id

    response = client.get(f"/dossiers/devis/{devis_id}/pdf")

    assert response.status_code == 200
    assert response.mimetype == "application/pdf"
    assert response.headers["Content-Disposition"].startswith("inline;")
    assert response.data.startswith(b"%PDF")
    assert b"MONSTER GARAGE" in response.data
    assert b"SNTL" not in response.data

    download = client.get(f"/dossiers/devis/{devis_id}/pdf/telecharger")

    assert download.status_code == 200
    assert download.mimetype == "application/pdf"
    assert download.headers["Content-Disposition"].startswith("attachment;")
    assert ".pdf" in download.headers["Content-Disposition"]


def test_export_facture_particulier_excel(client, app):
    connecter(client)
    dossier_id = creer_dossier_termine(client, app)
    client.post(f"/factures/dossiers/{dossier_id}/generer")
    with app.app_context():
        facture_id = FactureReparation.query.first().id

    response = client.get(f"/factures/{facture_id}/telecharger")

    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.data), data_only=False)
    assert_classeur_sans_sntl(wb)
    ws = wb.active
    assert ws["C4"].value == "FACTURE"
    assert ws["G4"].value.startswith("FA-")
    assert ws.freeze_panes is None
    assert ws["E22"].number_format == "0.##"
    assert ws["G24"].value == 200
    assert ws["G25"].value == 40
    assert ws["G26"].value == 240
    assert ws["G27"].value == 0
    assert ws["G28"].value == 240
    assert any(cell.value == "Montant TTC" for row in ws.iter_rows() for cell in row)


def test_export_facture_pdf_vue_et_telechargement(client, app):
    connecter(client)
    dossier_id = creer_dossier_termine(client, app)
    client.post(f"/factures/dossiers/{dossier_id}/generer")
    with app.app_context():
        facture_id = FactureReparation.query.first().id

    detail = client.get(f"/factures/{facture_id}")
    assert detail.status_code == 200
    assert b"Exporter Excel" in detail.data
    assert b"Voir PDF" in detail.data

    response = client.get(f"/factures/{facture_id}/pdf")

    assert response.status_code == 200
    assert response.mimetype == "application/pdf"
    assert response.headers["Content-Disposition"].startswith("inline;")
    assert response.data.startswith(b"%PDF")
    assert b"FACTURE" in response.data
    assert b"SNTL" not in response.data

    download = client.get(f"/factures/{facture_id}/pdf/telecharger")

    assert download.status_code == 200
    assert download.mimetype == "application/pdf"
    assert download.headers["Content-Disposition"].startswith("attachment;")
    assert ".pdf" in download.headers["Content-Disposition"]


def test_export_facture_sntl_ajoute_commission(client, app):
    connecter(client)
    with app.app_context():
        client_db = Client(
            code="SNTL1",
            type="sntl",
            nom="SNTL Marrakech",
            ice="ICE-SNTL-001",
            administration_rattachee="COMMUNE HARBIL",
        )
        db.session.add(client_db)
        db.session.flush()
        vehicule = Vehicule(
            client_id=client_db.id,
            immatriculation="J207789",
            marque="Ford",
            modele="Transit",
            kilometrage_actuel=332988,
        )
        db.session.add(vehicule)
        db.session.commit()
        client_id = client_db.id
        vehicule_id = vehicule.id

    client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "existant",
            "client_id": str(client_id),
            "vehicule_id": str(vehicule_id),
            "demande_client": "Freinage",
            "diagnostic_initial": "",
            "kilometrage_entree": "333000",
            "numero_bon_sntl": "260429000021",
            "notes": "",
        },
    )
    with app.app_context():
        dossier_id = DossierReparation.query.first().id
    client.post(
        f"/dossiers/{dossier_id}/devis/nouveau",
        data={
            "objet": "Freinage",
            "designation_1": "PLAQUETTE DE FREIN",
            "quantite_1": "1",
            "prix_unitaire_ht_1": "874.50",
            "designation_2": "Main d'oeuvre",
            "quantite_2": "1",
            "prix_unitaire_ht_2": "125.50",
        },
    )
    with app.app_context():
        devis_id = db.session.get(DossierReparation, dossier_id).dernier_devis.id
    devis_response = client.get(f"/dossiers/devis/{devis_id}/telecharger")
    devis_ws = load_workbook(BytesIO(devis_response.data), data_only=False).active
    devis_values = [cell.value for row in devis_ws.iter_rows() for cell in row if cell.value is not None]
    assert devis_ws["A1"].value == "MONSTER GARAGE"
    assert devis_ws["D11"].value == "Devis N°"
    assert devis_ws["E11"].value.endswith("-V1")
    assert devis_ws["C15"].value == "La Société Nationale des Transports et de la Logistique (SNTL)"
    assert "ETAT" not in devis_values
    assert devis_ws["B18"].value == "Partenaire"
    assert devis_ws["D18"].value == "Véhicule"
    assert devis_ws["C26"].value == "003524622000063"
    assert devis_ws["E19"].value == "0207789 - J"
    assert devis_ws["E23"].value == "260429000021"
    assert devis_ws["E31"].value == "Quantité"
    assert devis_ws["E32"].value == 1
    assert devis_ws["C32"].value == "PLAQUETTE DE FREIN"
    assert devis_ws["B42"].value == "Montant Total article HT"
    assert devis_ws["G42"].value == 874.5
    assert devis_ws["B43"].value == "Main d'œuvre*"
    assert devis_ws["G43"].value == 125.5
    devis_pdf = client.get(f"/dossiers/devis/{devis_id}/pdf")
    assert devis_pdf.status_code == 200
    assert devis_pdf.data.startswith(b"%PDF")
    assert b"Devis N" in devis_pdf.data
    assert b"SNTL" in devis_pdf.data
    assert b"Partenaire" in devis_pdf.data
    assert b"Matricule" in devis_pdf.data
    assert b"0207789 - J" in devis_pdf.data
    assert b"ETAT" not in devis_pdf.data
    assert b"Commission SNTL" not in devis_pdf.data
    client.post(f"/dossiers/devis/{devis_id}/approuver", data={"mode_accord": "telephone"})
    client.post(f"/dossiers/{dossier_id}/terminer")
    client.post(f"/factures/dossiers/{dossier_id}/generer")
    with app.app_context():
        facture_id = FactureReparation.query.first().id

    response = client.get(f"/factures/{facture_id}/telecharger")

    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.data), data_only=False)
    ws = wb.active
    assert ws["A1"].value == "MONSTER GARAGE"
    assert ws["D11"].value == "Facture N°"
    assert ws["E11"].value.startswith("FA-")
    assert ws["F11"].value == "Date facture:"
    assert ws["C15"].value == "La Société Nationale des Transports et de la Logistique (SNTL)"
    assert ws["B18"].value == "Partenaire"
    assert ws["D18"].value == "Véhicule"
    assert ws["B19"].value == "N° Agrément SNTL"
    assert ws["C19"].value == "3108"
    assert ws["B26"].value == "ICE"
    assert ws["C26"].value == "003524622000063"
    assert ws["D19"].value == "Matricule"
    assert ws["E19"].value == "0207789 - J"
    assert ws["D20"].value == "Marque et modèle"
    assert ws["E20"].value == "FORD TRANSIT"
    assert ws["D22"].value == "Administration"
    assert ws["E22"].value == "COMMUNE HARBIL"
    assert ws["E23"].value == "260429000021"
    assert ws.freeze_panes is None
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert "ETAT" not in values
    assert "ICE-SNTL-001" not in values
    assert ws["B31"].value == "Référence article"
    assert ws["C31"].value == "Désignation Article"
    assert ws["E31"].value == "Quantité"
    assert ws["F31"].value == "PU HT"
    assert ws["G31"].value == "Total HT"
    assert ws["B32"].value == "REF - N001"
    assert ws["C32"].value == "PLAQUETTE DE FREIN"
    assert ws["E32"].value == 1
    assert ws["F32"].value == 874.5
    assert ws["G32"].value == 874.5
    assert ws["B43"].value == "Main d'œuvre*"
    assert ws["G43"].value == 125.5
    assert ws["G44"].value == 1000
    assert ws["G45"].value == 200
    assert ws["G46"].value == 1200
    assert ws["G47"].value == 100
    assert ws["G48"].value == 20
    assert ws["G49"].value == 1080
    assert "Commission SNTL (1x10%) (4)" in values
    assert "Montant Net à régler (3-4-5)" in values
    assert ws["B51"].value.startswith("Arrêté la présente facture")
    assert ws["E54"].value == "Cachet et signature"
    assert 874.5 in values
    assert 200 in values
    facture_pdf = client.get(f"/factures/{facture_id}/pdf")
    assert facture_pdf.status_code == 200
    assert facture_pdf.data.startswith(b"%PDF")
    assert b"Facture N" in facture_pdf.data
    assert b"SNTL" in facture_pdf.data
    assert b"Partenaire" in facture_pdf.data
    assert b"Matricule" in facture_pdf.data
    assert b"0207789 - J" in facture_pdf.data
    assert b"Commission SNTL" in facture_pdf.data
    assert b"Montant Net" in facture_pdf.data
    assert b"ETAT" not in facture_pdf.data


def test_export_releve_client_excel(client, app):
    connecter(client)
    dossier_id = creer_dossier_termine(client, app)
    client.post(f"/factures/dossiers/{dossier_id}/generer")
    with app.app_context():
        facture = FactureReparation.query.first()
        client_id = facture.dossier.client_id

    client.post(
        "/dossiers/nouveau",
        data={
            "mode_client": "existant",
            "mode_vehicule": "nouveau",
            "client_id": str(client_id),
            "vehicule_immatriculation": "2222-A-6",
            "vehicule_marque": "Renault",
            "vehicule_modele": "Master",
            "demande_client": "Deuxieme vehicule",
            "diagnostic_initial": "",
            "kilometrage_entree": "70000",
            "notes": "",
        },
    )
    with app.app_context():
        second_dossier_id = DossierReparation.query.order_by(DossierReparation.id.desc()).first().id
    client.post(
        f"/dossiers/{second_dossier_id}/devis/nouveau",
        data={"objet": "Deuxieme facture", "designation_1": "Revision", "quantite_1": "1", "prix_unitaire_ht_1": "300"},
    )
    with app.app_context():
        second_devis_id = db.session.get(DossierReparation, second_dossier_id).dernier_devis.id
    client.post(f"/dossiers/devis/{second_devis_id}/approuver", data={"mode_accord": "telephone"})
    client.post(f"/dossiers/{second_dossier_id}/terminer")
    client.post(f"/factures/dossiers/{second_dossier_id}/generer")

    response = client.get(f"/factures/clients/{client_id}/releve")

    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.data), data_only=False)
    assert_classeur_sans_sntl(wb)
    ws = wb.active
    assert ws["A9"].value == "N FACTURE"
    assert ws["D9"].value == "MONTANT FACTURE"
    assert ws.freeze_panes is None
    assert ws["F10"].value == "=D10-E10"
    assert any(cell.value == "TOTAL" for row in ws.iter_rows() for cell in row)
    assert "RESUME VEHICULES" in wb.sheetnames
    resume = wb["RESUME VEHICULES"]
    assert resume["A7"].value == "VEHICULE"
    assert resume["B7"].value == "IMMATRICULATION"
    assert resume["A8"].value == "Toyota Hilux"
    assert resume["B8"].value == "1111-F-6"
    assert resume["C8"].value == 1
    assert resume["D8"].value == 240
    assert resume["F8"].value == 240
    assert resume["A9"].value == "Renault Master"
    assert resume["B9"].value == "2222-A-6"
    assert resume["C9"].value == 1
    assert resume["D9"].value == 360
    assert resume["F9"].value == 360


def _classeur_salaires(lignes):
    wb = Workbook()
    ws = wb.active
    ws.title = "SALAIRES"
    ws.append(["Nom & Prénom", "Date", "Montant", "Fonction", "Observations"])
    for ligne in lignes:
        ws.append(ligne)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_import_rh_ne_duplique_pas_operations_non_salariales(app):
    data = _classeur_salaires([
        ["ALI", date(2026, 4, 10), 100, "OUVRIER", "AVANCE"],
        ["ALI", date(2026, 4, 12), 300, "OUVRIER", "PRIME"],
    ])

    with app.app_context():
        resultat_1 = importer_salaires_excel(BytesIO(data), mois=4, annee=2026)
        resultat_2 = importer_salaires_excel(BytesIO(data), mois=4, annee=2026)

        assert resultat_1["avances_creees"] == 2
        assert resultat_2["avances_creees"] == 0
        assert len(resultat_2["ignores"]) == 2
        assert AvanceSalaire.query.count() == 2


def test_import_rh_fallback_date_utilise_dernier_jour_du_mois(app):
    data = _classeur_salaires([
        ["ALI", None, 150, "OUVRIER", "AVANCE"],
    ])

    with app.app_context():
        resultat = importer_salaires_excel(BytesIO(data), mois=2, annee=2026)
        avance = AvanceSalaire.query.one()

        assert resultat["avances_creees"] == 1
        assert avance.date == date(2026, 2, 28)


def test_import_rh_ligne_quinzaine_renseigne_salaire_quinzaine(app):
    data = _classeur_salaires([
        ["ABDESLAM", date(2026, 4, 15), 2250, "TOLIER", "QUINZ 15/04/2026"],
    ])

    with app.app_context():
        importer_salaires_excel(BytesIO(data), mois=4, annee=2026)
        employe = Employe.query.filter_by(nom_complet="ABDESLAM").one()

        assert str(employe.salaire_quinzaine) == "2250.00"
        assert str(employe.salaires[0].montant_net_paye) == "2250.00"


def test_rh_refuse_salarie_fixe_sans_salaire_quinzaine(client, app):
    connecter(client)

    response = client.post(
        "/rh/employes/nouveau",
        data={
            "nom_complet": "EMPLOYE FIXE",
            "fonction": "ouvrier",
            "type_remuneration": "salaire_fixe",
            "salaire_quinzaine": "",
        },
    )

    assert response.status_code == 200
    with app.app_context():
        assert Employe.query.filter_by(nom_complet="EMPLOYE FIXE").first() is None


def test_rh_refuse_type_operation_invalide(client, app):
    connecter(client)
    with app.app_context():
        employe = Employe(nom_complet="ALI", fonction="ouvrier", type_remuneration="tache", actif=True)
        db.session.add(employe)
        db.session.commit()
        employe_id = employe.id

    response = client.post(
        "/rh/avances/nouvelle",
        data={
            "employe_id": str(employe_id),
            "date": "2026-04-10",
            "montant": "100",
            "type": "bonus_inconnu",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert response.status_code == 400
    with app.app_context():
        assert AvanceSalaire.query.count() == 0


def test_rh_recap_totalise_reste_du_de_facon_coherente(app):
    from app.services.calcul_salaire import get_alertes_fin_mois, get_recap_mensuel, get_total_mensuel

    with app.app_context():
        employe = Employe(nom_complet="ALI", fonction="ouvrier", type_remuneration="tache", actif=True)
        db.session.add(employe)
        db.session.flush()
        db.session.add_all([
            AvanceSalaire(
                employe_id=employe.id,
                date=date(2026, 4, 10),
                montant=1000,
                type="tache",
                description="Mission partielle",
                quinzaine="premiere",
                mois=4,
                annee=2026,
                montant_total_convenu=1500,
                reste_du=500,
            ),
            AvanceSalaire(
                employe_id=employe.id,
                date=date(2026, 4, 20),
                montant=200,
                type="reste_du",
                description="Paiement partiel reste",
                quinzaine="seconde",
                mois=4,
                annee=2026,
                montant_total_convenu=500,
                reste_du=300,
            ),
        ])
        db.session.commit()

        ligne = get_recap_mensuel(4, 2026)[0]
        alertes = get_alertes_fin_mois(4, 2026)

        assert str(ligne["total_primes"]) == "1200.00"
        assert str(ligne["reste_du"]) == "800.00"
        assert str(ligne["total_mois"]) == str(get_total_mensuel(4, 2026))
        assert len([alerte for alerte in alertes if alerte["type"] == "reste_du"]) == 2


def test_rh_avance_ne_augmente_pas_le_net_salaire(app):
    from app.services.calcul_salaire import enregistrer_quinzaine, enregistrer_solde_fin_mois, get_recap_mensuel

    with app.app_context():
        employe = Employe(
            nom_complet="ABDESLAM",
            fonction="tolier",
            type_remuneration="salaire_fixe",
            salaire_quinzaine=2250,
            actif=True,
        )
        db.session.add(employe)
        db.session.flush()
        db.session.add_all([
            AvanceSalaire(
                employe_id=employe.id,
                date=date(2026, 5, 10),
                montant=500,
                type="avance",
                quinzaine="premiere",
                mois=5,
                annee=2026,
            ),
            AvanceSalaire(
                employe_id=employe.id,
                date=date(2026, 5, 24),
                montant=600,
                type="avance",
                quinzaine="seconde",
                mois=5,
                annee=2026,
            ),
        ])
        enregistrer_quinzaine(employe, 5, 2026)
        db.session.commit()

        ligne = get_recap_mensuel(5, 2026)[0]
        assert str(ligne["total_avances"]) == "1100.00"
        assert str(ligne["total_mois"]) == "2250.00"

        enregistrer_solde_fin_mois(employe, 5, 2026)
        db.session.commit()

        ligne = get_recap_mensuel(5, 2026)[0]
        assert str(ligne["solde"].total_avances) == "600.00"
        assert str(ligne["solde"].montant_net_paye) == "1650.00"
        assert str(ligne["total_mois"]) == "3900.00"


def test_rh_solde_accepte_brut_solde_exceptionnel(app):
    from app.services.calcul_salaire import enregistrer_solde_fin_mois

    with app.app_context():
        employe = Employe(
            nom_complet="SAID",
            fonction="chef_atelier",
            type_remuneration="mixte",
            salaire_quinzaine=3700,
            actif=True,
        )
        db.session.add(employe)
        db.session.flush()
        for jour in (17, 22, 27):
            db.session.add(
                AvanceSalaire(
                    employe_id=employe.id,
                    date=date(2026, 4, jour),
                    montant=200,
                    type="avance",
                    quinzaine="seconde",
                    mois=4,
                    annee=2026,
                )
            )

        salaire = enregistrer_solde_fin_mois(employe, 4, 2026, brut_solde=5000)
        db.session.commit()

        assert str(salaire.salaire_brut) == "5000.00"
        assert str(salaire.total_avances) == "600.00"
        assert str(salaire.montant_net_paye) == "4400.00"


def test_rh_peut_regler_quinzaine_un_seul_employe(client, app):
    connecter(client)
    with app.app_context():
        abd = Employe(
            nom_complet="ABDESLAM",
            fonction="tolier",
            type_remuneration="salaire_fixe",
            salaire_quinzaine=2250,
            actif=True,
        )
        said = Employe(
            nom_complet="SAID",
            fonction="chef_atelier",
            type_remuneration="mixte",
            salaire_quinzaine=3700,
            actif=True,
        )
        db.session.add_all([abd, said])
        db.session.commit()
        abd_id = abd.id

    response = client.post(
        f"/rh/employes/{abd_id}/payer-quinzaine",
        data={"mois": "4", "annee": "2026"},
    )

    assert response.status_code == 302
    with app.app_context():
        assert Salaire.query.count() == 1
        salaire = Salaire.query.one()
        assert salaire.employe.nom_complet == "ABDESLAM"
        assert salaire.type_paie == "quinzaine"
        assert str(salaire.montant_net_paye) == "2250.00"


def test_rh_peut_calculer_solde_un_seul_employe(client, app):
    connecter(client)
    with app.app_context():
        abd = Employe(
            nom_complet="ABDESLAM",
            fonction="tolier",
            type_remuneration="salaire_fixe",
            salaire_quinzaine=2250,
            actif=True,
        )
        said = Employe(
            nom_complet="SAID",
            fonction="chef_atelier",
            type_remuneration="mixte",
            salaire_quinzaine=3700,
            actif=True,
        )
        db.session.add_all([abd, said])
        db.session.flush()
        db.session.add(
            AvanceSalaire(
                employe_id=abd.id,
                date=date(2026, 4, 24),
                montant=600,
                type="avance",
                quinzaine="seconde",
                mois=4,
                annee=2026,
            )
        )
        db.session.commit()
        abd_id = abd.id

    response = client.post(
        f"/rh/employes/{abd_id}/calculer-solde",
        data={"mois": "4", "annee": "2026", "brut_solde": "2250"},
    )

    assert response.status_code == 302
    with app.app_context():
        assert Salaire.query.count() == 1
        salaire = Salaire.query.one()
        assert salaire.employe.nom_complet == "ABDESLAM"
        assert salaire.type_paie == "fin_mois"
        assert str(salaire.total_avances) == "600.00"
        assert str(salaire.montant_net_paye) == "1650.00"


def test_export_rh_contient_logo_et_grand_titre_atelier(app):
    with app.app_context():
        data = exporter_salaires_excel(5, 2026)

    wb = load_workbook(BytesIO(data))
    ws = wb["SALAIRES"]

    assert ws["C1"].value == "MONSTER GARAGE"
    assert ws["C1"].font.sz == 24
    assert "LIVRE DE PAIE" in ws["C4"].value
    assert "MAI 2026" in ws["C4"].value
    assert len(ws._images) == 1
